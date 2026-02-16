
import asyncio
import openpyxl
from pathlib import Path
from sqlalchemy import select
from app.core.config import get_settings
from app.db.session import session_context
from app.models import Risk

async def migrate_risk_names_from_excel():
    excel_path = Path("Registr_Rizik_2022.xlsx")
    if not excel_path.exists():
        print(f"❌ Excel file not found: {excel_path}")
        return
    
    print(f"📂 Loading: {excel_path}")
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    sheet = wb['Rizika']
    
    # Map from risk_id_code to (name, description)
    # We'll re-generate the codes exactly as in the original migration script
    # to match them back correctly.
    excel_data = {}
    from collections import defaultdict
    process_counters = defaultdict(int)
    
    def get_process_code(process_name):
        if not process_name: return 'UNK'
        clean = process_name.upper()[:3]
        replacements = {'Á': 'A', 'É': 'E', 'Í': 'I', 'Ó': 'O', 'Ú': 'U', 'Ý': 'Y', 'Č': 'C', 'Š': 'S', 'Ž': 'Z', 'Ř': 'R', 'Ň': 'N', 'Ť': 'T', 'Ď': 'D'}
        for k, v in replacements.items():
            clean = clean.replace(k, v)
        return clean

    print("🔍 Parsing Excel data...")
    for row_idx in range(9, sheet.max_row + 1):
        row = list(sheet.iter_rows(min_row=row_idx, max_row=row_idx, values_only=True))[0]
        hlavni_proces = str(row[1]).strip() if row[1] else None
        
        # Original migration skipped rows without process or description (Col F)
        popis_f = str(row[5]).strip() if row[5] else None # Column F
        
        if not hlavni_proces or not popis_f:
            continue
            
        proc_code = get_process_code(hlavni_proces)
        process_counters[proc_code] += 1
        risk_id_code = f"{proc_code}-R{process_counters[proc_code]:02d}"
        
        name = popis_f
        description = str(row[6]).strip() if row[6] else "" # Column G
        
        excel_data[risk_id_code] = (name, description)

    print(f"🚀 Updating Risks in Database...")
    async with session_context(get_settings()) as session:
        result = await session.execute(select(Risk))
        risks = result.scalars().all()
        
        updated_count = 0
        for risk in risks:
            if risk.risk_id_code in excel_data:
                new_name, new_desc = excel_data[risk.risk_id_code]
                
                # Update if changed
                if risk.name != new_name or risk.description != new_desc:
                    print(f"🔄 Updating {risk.risk_id_code}:")
                    if risk.name != new_name:
                        print(f"   Name: '{risk.name}' -> '{new_name}'")
                    if risk.description != new_desc:
                        print(f"   Desc: '{risk.description[:50]}...' -> '{new_desc[:50]}...'")
                    
                    risk.name = new_name
                    risk.description = new_desc
                    updated_count += 1
            else:
                print(f"⚠️  No Excel data found for {risk.risk_id_code}")
        
        if updated_count > 0:
            await session.commit()
            print(f"\n✅ Successfully updated {updated_count} risks.")
        else:
            print("\nℹ️ No risks required updates.")

if __name__ == "__main__":
    asyncio.run(migrate_risk_names_from_excel())
