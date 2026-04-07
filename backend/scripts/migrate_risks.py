#!/usr/bin/env python3
"""Deterministic risk import with safe non-reset identity matching."""

from __future__ import annotations

import argparse
import asyncio
import sys
from collections import defaultdict
from dataclasses import dataclass
from pathlib import Path

import openpyxl
from sqlalchemy import delete, select

from app.api.v1.endpoints.risks.id_generation import generate_risk_id_code
from app.core.config import get_settings
from app.db.session import session_context
from app.models import ControlRiskLink, Department, KeyRiskIndicator, Risk, User
from scripts.import_contracts import ImportReport, write_report

SHEET_NAME = "Rizika"


@dataclass(slots=True)
class RiskImportRow:
    row_number: int
    process: str
    subprocess: str | None
    name: str
    description: str
    risk_type: str
    category: str
    gross_impact: int
    gross_probability: int
    net_impact: int
    net_probability: int


@dataclass(slots=True)
class DepartmentPlan:
    process_name: str
    code: str
    existing: Department | None


def _parse_args(argv: list[str]) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Import risks from an Excel workbook.")
    parser.add_argument("--input", required=True, help="Path to the source workbook.")
    parser.add_argument("--apply", action="store_true", help="Persist changes. Default is dry-run.")
    parser.add_argument(
        "--allow-reset",
        action="store_true",
        help="Allow destructive reset of existing risks, KRIs, and control-risk links before import.",
    )
    parser.add_argument("--report", help="Optional path for a JSON import report.")
    return parser.parse_args(argv)


def _normalize_string(value: object | None) -> str:
    if value is None:
        return ""
    return " ".join(str(value).replace("\n", " ").split()).lower().strip()


def _identity_key(process: object | None, subprocess: object | None, name: object | None) -> tuple[str, str, str]:
    return (
        _normalize_string(process),
        _normalize_string(subprocess),
        _normalize_string(name),
    )


def _process_key(process_name: object | None) -> str:
    return _normalize_string(process_name)


def _safe_int(value: object | None, default: int) -> int:
    try:
        return int(value) if value is not None and str(value).strip() else default
    except (TypeError, ValueError):
        return default


def _clamp_score(value: int) -> int:
    return max(1, min(5, value))


def _get_process_code(process_name: str) -> str:
    clean = process_name.upper()[:3]
    replacements = {
        "Á": "A",
        "É": "E",
        "Í": "I",
        "Ó": "O",
        "Ú": "U",
        "Ý": "Y",
        "Č": "C",
        "Š": "S",
        "Ž": "Z",
        "Ř": "R",
        "Ň": "N",
        "Ť": "T",
        "Ď": "D",
    }
    for source, target in replacements.items():
        clean = clean.replace(source, target)
    return clean or "UNK"


def _load_rows(input_path: Path, report: ImportReport) -> list[RiskImportRow]:
    workbook = openpyxl.load_workbook(input_path, data_only=True)
    if SHEET_NAME not in workbook.sheetnames:
        report.add_error("missing-sheet", f"Workbook does not contain required sheet {SHEET_NAME!r}.")
        return []

    sheet = workbook[SHEET_NAME]
    rows: list[RiskImportRow] = []
    seen_workbook_keys: dict[tuple[str, str, str], int] = {}
    duplicate_workbook_keys: list[dict[str, object]] = []

    for row_idx in range(9, sheet.max_row + 1):
        values = list(sheet.iter_rows(min_row=row_idx, max_row=row_idx, values_only=True))[0]
        process = str(values[1]).strip() if values[1] else ""
        subprocess = str(values[2]).strip() if values[2] else None
        name = str(values[5]).strip() if values[5] else ""
        description = str(values[6]).strip() if values[6] else ""

        if not process or not name:
            report.skips += 1
            continue

        row_key = _identity_key(process, subprocess, name)
        if row_key in seen_workbook_keys:
            duplicate = {
                "process": process,
                "subprocess": subprocess,
                "name": name,
                "first_row": seen_workbook_keys[row_key],
                "second_row": row_idx,
            }
            duplicate_workbook_keys.append(duplicate)
            report.add_error(
                "duplicate-workbook-key",
                (
                    f"Workbook contains duplicate risk identity for process={process!r}, "
                    f"subprocess={subprocess!r}, name={name!r}."
                ),
                sheet=SHEET_NAME,
                row=row_idx,
                details=duplicate,
            )
            continue
        seen_workbook_keys[row_key] = row_idx

        source_risk_type = str(values[3]).strip() if values[3] else "Operační riziko"
        risk_type = "strategic" if "strateg" in _normalize_string(source_risk_type) else "operational"
        category = source_risk_type

        rows.append(
            RiskImportRow(
                row_number=row_idx,
                process=process,
                subprocess=subprocess,
                name=name,
                description=description,
                risk_type=risk_type,
                category=category,
                gross_impact=_clamp_score(_safe_int(values[7], 3)),
                gross_probability=_clamp_score(_safe_int(values[8], 3)),
                net_impact=_clamp_score(_safe_int(values[11], 2)),
                net_probability=_clamp_score(_safe_int(values[12], 2)),
            )
        )

    report.metadata["rows_loaded"] = len(rows)
    report.metadata["duplicate_workbook_keys"] = duplicate_workbook_keys
    if not rows and report.ok:
        report.add_error("no-rows", "Workbook did not contain any importable risk rows.", sheet=SHEET_NAME)
    return rows


def _plan_departments(rows: list[RiskImportRow], existing_departments: list[Department]) -> dict[str, DepartmentPlan]:
    departments_by_name = {_process_key(department.name): department for department in existing_departments}
    used_codes = {department.code.upper() for department in existing_departments}
    plans: dict[str, DepartmentPlan] = {}

    for row in rows:
        process_key = _process_key(row.process)
        if process_key in plans:
            continue

        existing = departments_by_name.get(process_key)
        if existing is not None:
            plans[process_key] = DepartmentPlan(process_name=existing.name, code=existing.code, existing=existing)
            continue

        base_code = _get_process_code(row.process)
        candidate = base_code
        suffix = 1
        while candidate.upper() in used_codes:
            candidate = f"{base_code}{suffix}"
            suffix += 1
        used_codes.add(candidate.upper())
        plans[process_key] = DepartmentPlan(process_name=row.process, code=candidate, existing=None)

    return plans


def _risk_payload(
    *,
    row: RiskImportRow,
    department_id: int,
    owner_id: int | None,
    risk_id_code: str | None = None,
) -> dict[str, object]:
    payload: dict[str, object] = {
        "name": row.name[:255],
        "process": row.process,
        "subprocess": row.subprocess,
        "risk_type": row.risk_type,
        "category": row.category,
        "description": row.description,
        "department_id": department_id,
        "owner_id": owner_id,
        "gross_probability": row.gross_probability,
        "gross_impact": row.gross_impact,
        "gross_score": row.gross_probability * row.gross_impact,
        "net_probability": row.net_probability,
        "net_impact": row.net_impact,
        "net_score": row.net_probability * row.net_impact,
        "status": "active",
        "is_priority": (row.gross_probability * row.gross_impact) >= 15,
    }
    if risk_id_code is not None:
        payload["risk_id_code"] = risk_id_code
    return payload


async def _run(args: argparse.Namespace) -> int:
    input_path = Path(args.input).expanduser().resolve()
    report = ImportReport(
        script="migrate_risks.py",
        input_path=str(input_path),
        apply=args.apply,
        allow_reset=args.allow_reset,
    )

    if not input_path.exists():
        report.add_error("missing-input", f"Input workbook not found: {input_path}")
        write_report(args.report, report)
        print(f"❌ Input workbook not found: {input_path}")
        return 1

    rows = _load_rows(input_path, report)

    async with session_context(get_settings()) as session:
        owner = (await session.execute(select(User).order_by(User.id).limit(1))).scalar_one_or_none()
        if owner is None:
            report.add_error("missing-owner", "No users found. Seed base data before importing risks.")

        existing_departments = list((await session.execute(select(Department))).scalars().all())
        department_plans = _plan_departments(rows, existing_departments)
        report.metadata["departments_to_create"] = [
            {"name": plan.process_name, "code": plan.code}
            for plan in department_plans.values()
            if plan.existing is None
        ]

        existing_risks = list((await session.execute(select(Risk))).scalars().all())
        matched_rows: list[tuple[RiskImportRow, Risk]] = []
        new_rows: list[RiskImportRow] = []

        if args.allow_reset:
            report.creates = len(rows)
            report.updates = 0
            report.deletes = len(existing_risks)
        else:
            report.metadata["identity_mode"] = "process_subprocess_name"
            existing_by_identity: defaultdict[tuple[str, str, str], list[Risk]] = defaultdict(list)
            for risk in existing_risks:
                existing_by_identity[_identity_key(risk.process, risk.subprocess, risk.name)].append(risk)

            ambiguous_matches: list[dict[str, object]] = []

            for row in rows:
                matches = existing_by_identity.get(_identity_key(row.process, row.subprocess, row.name), [])
                if len(matches) == 1:
                    matched_rows.append((row, matches[0]))
                    continue
                if len(matches) == 0:
                    new_rows.append(row)
                    continue

                ambiguous = {
                    "row": row.row_number,
                    "process": row.process,
                    "subprocess": row.subprocess,
                    "name": row.name,
                    "matched_risk_ids": [risk.id for risk in matches],
                    "matched_risk_codes": [risk.risk_id_code for risk in matches],
                }
                ambiguous_matches.append(ambiguous)
                report.add_error(
                    "ambiguous-existing-risk-match",
                    (
                        f"Workbook row {row.row_number} matched multiple existing risks for "
                        f"process={row.process!r}, subprocess={row.subprocess!r}, name={row.name!r}."
                    ),
                    sheet=SHEET_NAME,
                    row=row.row_number,
                    details=ambiguous,
                )

            report.creates = len(new_rows)
            report.updates = len(matched_rows)
            report.metadata["matched_count"] = len(matched_rows)
            report.metadata["created_count"] = len(new_rows)
            report.metadata["ambiguous_matches"] = ambiguous_matches

        if not report.ok:
            write_report(args.report, report)
            for issue in report.errors:
                print(f"❌ {issue.code}: {issue.message}")
            return 1

        if not args.apply:
            write_report(args.report, report)
            print("🔍 DRY-RUN: no changes applied.")
            if args.allow_reset:
                print(
                    f"   creates={report.creates} updates={report.updates} deletes={report.deletes} "
                    f"skips={report.skips} dept_creates={len(report.metadata['departments_to_create'])}"
                )
            else:
                print(
                    f"   creates={report.creates} updates={report.updates} deletes={report.deletes} "
                    f"skips={report.skips} matched={report.metadata['matched_count']}"
                )
            return 0

        if args.allow_reset:
            await session.execute(delete(ControlRiskLink))
            await session.execute(delete(KeyRiskIndicator))
            await session.execute(delete(Risk))
            await session.flush()

        departments_by_process: dict[str, Department] = {}
        for process_key, plan in department_plans.items():
            if plan.existing is not None:
                departments_by_process[process_key] = plan.existing
                continue
            department = Department(
                name=plan.process_name,
                code=plan.code,
                description=f"Department for imported process {plan.process_name}",
            )
            session.add(department)
            await session.flush()
            departments_by_process[process_key] = department

        if args.allow_reset:
            for row in rows:
                department = departments_by_process[_process_key(row.process)]
                risk_id_code = await generate_risk_id_code(session, row.process)
                session.add(
                    Risk(
                        **_risk_payload(
                            row=row,
                            department_id=department.id,
                            owner_id=owner.id if owner is not None else None,
                            risk_id_code=risk_id_code,
                        )
                    )
                )
                await session.flush()
            await session.commit()
        else:
            for row, risk in matched_rows:
                department = departments_by_process[_process_key(row.process)]
                payload = _risk_payload(
                    row=row,
                    department_id=department.id,
                    owner_id=owner.id if owner is not None else None,
                )
                for field_name, field_value in payload.items():
                    setattr(risk, field_name, field_value)

            for row in new_rows:
                department = departments_by_process[_process_key(row.process)]
                risk_id_code = await generate_risk_id_code(session, row.process)
                risk = Risk(
                    **_risk_payload(
                        row=row,
                        department_id=department.id,
                        owner_id=owner.id if owner is not None else None,
                        risk_id_code=risk_id_code,
                    )
                )
                session.add(risk)
                await session.flush()

            await session.commit()

    write_report(args.report, report)
    print(
        f"✅ Imported risks: creates={report.creates} updates={report.updates} "
        f"deletes={report.deletes} skips={report.skips}"
    )
    return 0


def main() -> int:
    return asyncio.run(_run(_parse_args(sys.argv[1:])))


if __name__ == "__main__":
    raise SystemExit(main())
