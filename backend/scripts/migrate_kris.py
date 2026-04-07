#!/usr/bin/env python3
"""Deterministic KRI import with dry-run/report/apply/reset contracts."""

from __future__ import annotations

import argparse
import asyncio
import re
import sys
from dataclasses import dataclass
from difflib import SequenceMatcher
from pathlib import Path

import openpyxl
from sqlalchemy import delete, select

from app.core.config import get_settings
from app.db.session import session_context
from app.models import KeyRiskIndicator, Risk
from scripts.import_contracts import ImportReport, write_report

KRI_SHEETS = [
    "Provozní riziko",
    "Neživotní upisovací riziko",
    "Zdravotní upisovací riziko",
    "Tržní riziko",
    "Riziko selhání protistrany",
]


@dataclass(slots=True)
class KriImportRow:
    sheet: str
    row_number: int
    process: str
    risk_description: str
    metric_name: str
    current_value: float
    lower_limit: float
    upper_limit: float
    unit: str


def _parse_args(argv: list[str]) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Import KRIs from an Excel workbook.")
    parser.add_argument("--input", required=True, help="Path to the source workbook.")
    parser.add_argument("--apply", action="store_true", help="Persist changes. Default is dry-run.")
    parser.add_argument(
        "--allow-reset",
        action="store_true",
        help="Allow destructive reset of existing KRIs before import.",
    )
    parser.add_argument("--report", help="Optional path for a JSON import report.")
    return parser.parse_args(argv)


def _normalize_string(value: object | None) -> str:
    if value is None:
        return ""
    return " ".join(str(value).lower().strip().replace("\n", " ").split())


def _similarity_score(left: str, right: str) -> float:
    return SequenceMatcher(None, _normalize_string(left), _normalize_string(right)).ratio()


def _safe_float(value: object | None, default: float = 0.0) -> float:
    if value is None:
        return default
    if isinstance(value, (int, float)):
        return float(value)
    match = re.search(r"[-+]?\d*[\.,]?\d+", str(value))
    if match is None:
        return default
    return float(match.group().replace(",", "."))


def _unit_from_metric(metric_name: str) -> str:
    lowered = metric_name.lower()
    if "%" in metric_name or "procent" in lowered:
        return "%"
    if "dn" in lowered or "den" in lowered:
        return "days"
    if "počet" in lowered or "count" in lowered:
        return "count"
    return "value"


def _load_rows(input_path: Path, report: ImportReport) -> list[KriImportRow]:
    workbook = openpyxl.load_workbook(input_path, data_only=True)
    rows: list[KriImportRow] = []

    for sheet_name in KRI_SHEETS:
        if sheet_name not in workbook.sheetnames:
            report.add_warning("missing-sheet", f"Workbook is missing optional sheet {sheet_name!r}.", sheet=sheet_name)
            continue

        sheet = workbook[sheet_name]
        for row_idx in range(2, sheet.max_row + 1):
            metric_name = str(sheet.cell(row=row_idx, column=9).value or "").strip()
            if not metric_name:
                report.skips += 1
                continue

            rows.append(
                KriImportRow(
                    sheet=sheet_name,
                    row_number=row_idx,
                    process=str(sheet.cell(row=row_idx, column=2).value or "").strip(),
                    risk_description=str(sheet.cell(row=row_idx, column=6).value or "").strip(),
                    metric_name=metric_name,
                    current_value=_safe_float(sheet.cell(row=row_idx, column=10).value),
                    lower_limit=_safe_float(sheet.cell(row=row_idx, column=11).value),
                    upper_limit=_safe_float(sheet.cell(row=row_idx, column=12).value, 100.0),
                    unit=_unit_from_metric(metric_name),
                )
            )

    report.metadata["rows_loaded"] = len(rows)
    if not rows and report.ok:
        report.add_error("no-rows", "Workbook did not contain any importable KRI rows.")
    return rows


def _match_risk(row: KriImportRow, all_risks: list[Risk], risks_by_process: dict[str, list[Risk]]) -> Risk | None:
    scored: list[tuple[Risk, float]] = []
    for risk in all_risks:
        score = _similarity_score(row.risk_description, risk.description or "")
        if score > 0.4:
            scored.append((risk, score))
    if scored:
        scored.sort(key=lambda item: (-item[1], item[0].risk_id_code))
        return scored[0][0]

    process_key = _normalize_string(row.process)
    candidates = risks_by_process.get(process_key, [])
    if candidates:
        return sorted(candidates, key=lambda risk: risk.risk_id_code)[0]
    return None


async def _run(args: argparse.Namespace) -> int:
    input_path = Path(args.input).expanduser().resolve()
    report = ImportReport(
        script="migrate_kris.py",
        input_path=str(input_path),
        apply=args.apply,
        allow_reset=args.allow_reset,
    )

    if not input_path.exists():
        report.add_error("missing-input", f"Input workbook not found: {input_path}")
        write_report(args.report, report)
        print(f"❌ Input workbook not found: {input_path}")
        return 1

    rows = _load_rows(input_path, report)

    async with session_context(get_settings()) as session:
        all_risks = list((await session.execute(select(Risk))).scalars().all())
        if not all_risks:
            report.add_error("missing-risks", "No risks found. Import or seed risks before importing KRIs.")

        risks_by_process: dict[str, list[Risk]] = {}
        for risk in all_risks:
            risks_by_process.setdefault(_normalize_string(risk.process), []).append(risk)

        existing_kris = list((await session.execute(select(KeyRiskIndicator))).scalars().all())
        existing_by_key: dict[tuple[int, str], KeyRiskIndicator] = {}
        for kri in existing_kris:
            key = (kri.risk_id, _normalize_string(kri.metric_name))
            if key in existing_by_key:
                report.add_error(
                    "duplicate-existing-kri",
                    "Database already contains duplicate KRI metric names for the same risk. Resolve before importing.",
                    details={"risk_id": kri.risk_id, "metric_name": kri.metric_name},
                )
                continue
            existing_by_key[key] = kri

        matched_rows: list[tuple[KriImportRow, Risk]] = []
        planned_keys: dict[tuple[int, str], KriImportRow] = {}
        for row in rows:
            matched_risk = _match_risk(row, all_risks, risks_by_process)
            if matched_risk is None:
                report.add_error(
                    "unmatched-kri",
                    f"Unable to match metric {row.metric_name!r} to a risk.",
                    sheet=row.sheet,
                    row=row.row_number,
                    details={"process": row.process, "risk_description": row.risk_description},
                )
                continue

            key = (matched_risk.id, _normalize_string(row.metric_name))
            duplicate = planned_keys.get(key)
            if duplicate is not None:
                report.add_error(
                    "duplicate-kri-key",
                    f"Workbook would import duplicate metric {row.metric_name!r} for risk {matched_risk.risk_id_code}.",
                    sheet=row.sheet,
                    row=row.row_number,
                    details={"first_row": duplicate.row_number, "risk_id_code": matched_risk.risk_id_code},
                )
                continue

            planned_keys[key] = row
            matched_rows.append((row, matched_risk))

        if args.allow_reset:
            report.creates = len(matched_rows)
            report.updates = 0
            report.deletes = len(existing_kris)
        else:
            report.creates = sum(
                1 for row, risk in matched_rows if (risk.id, _normalize_string(row.metric_name)) not in existing_by_key
            )
            report.updates = len(matched_rows) - report.creates

        if not report.ok:
            write_report(args.report, report)
            for issue in report.errors:
                location = f" ({issue.sheet} row {issue.row})" if issue.sheet and issue.row else ""
                print(f"❌ {issue.code}{location}: {issue.message}")
            return 1

        if not args.apply:
            write_report(args.report, report)
            print("🔍 DRY-RUN: no changes applied.")
            print(
                f"   creates={report.creates} updates={report.updates} deletes={report.deletes} "
                f"skips={report.skips} warnings={len(report.warnings)}"
            )
            return 0

        if args.allow_reset:
            await session.execute(delete(KeyRiskIndicator))
            await session.flush()
            existing_by_key = {}

        for row, matched_risk in matched_rows:
            key = (matched_risk.id, _normalize_string(row.metric_name))
            kri = existing_by_key.get(key)
            description = row.risk_description or f"Imported from {row.sheet} row {row.row_number}"

            if kri is None:
                session.add(
                    KeyRiskIndicator(
                        risk_id=matched_risk.id,
                        metric_name=row.metric_name[:500],
                        description=description,
                        current_value=row.current_value,
                        lower_limit=row.lower_limit,
                        upper_limit=row.upper_limit,
                        unit=row.unit,
                    )
                )
                continue

            kri.metric_name = row.metric_name[:500]
            kri.description = description
            kri.current_value = row.current_value
            kri.lower_limit = row.lower_limit
            kri.upper_limit = row.upper_limit
            kri.unit = row.unit

        await session.commit()

    write_report(args.report, report)
    print(
        f"✅ Imported KRIs: creates={report.creates} updates={report.updates} "
        f"deletes={report.deletes} skips={report.skips}"
    )
    return 0


def main() -> int:
    return asyncio.run(_run(_parse_args(sys.argv[1:])))


if __name__ == "__main__":
    raise SystemExit(main())
