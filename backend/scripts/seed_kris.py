#!/usr/bin/env python3
"""
Seed KRI data from Excel spreadsheet with risk matching.

Usage:
    python seed_kris.py                     # Dry-run mode (log only)
    python seed_kris.py --force             # Actually delete and seed KRIs
    python seed_kris.py --report unmatched_kris.txt  # Write unmatched KRIs to file
"""
import argparse
import asyncio
import re
import sys
from pathlib import Path
from collections import defaultdict

# Mapping sheet names to database categories
SHEET_TO_CATEGORY = {
    "Neživotní upisovací riziko": "Upisovací riziko",
    "Zdravotní upisovací riziko": "Upisovací riziko",
    "Tržní riziko": "Tržní riziko",
    "Riziko selhání protistrany": "Selhání protistrany",
    "Provozní riziko": "Operační riziko",
}

# General keyword mapping for fallback
KEYWORD_TO_PROCESS = {
    "marketing": "Marketing",
    "it": "IT",
    "finance": "Finance",
    "hr": "Lidské zdroje",
    "personální": "Lidské zdroje",
    "likvidace": "Likvidace pojistných událostí",
}


def safe_float(val, default=0.0):
    if val is None:
        return default
    try:
        if isinstance(val, (int, float)):
            return float(val)
        match = re.search(r"[-+]?\d*[\.,]?\d+", str(val))
        if match:
            return float(match.group().replace(',', '.'))
        return default
    except Exception:
        return default


def parse_args():
    parser = argparse.ArgumentParser(
        description="Seed KRI data from Excel spreadsheet",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python seed_kris.py                          # Dry-run: preview what would happen
  python seed_kris.py --force                  # Actually delete existing KRIs and seed
  python seed_kris.py --force --report out.txt # Seed and write unmatched to file
        """
    )
    parser.add_argument(
        "--force",
        action="store_true",
        help="Actually perform destructive operations (delete existing KRIs). "
             "Without this flag, the script runs in dry-run mode."
    )
    parser.add_argument(
        "--report",
        type=str,
        metavar="FILE",
        help="Write unmatched KRIs to this file for manual review"
    )
    parser.add_argument(
        "--excel",
        type=str,
        metavar="PATH",
        help="Path to Excel file (default: placeholder-kri-source.xlsx)"
    )
    return parser.parse_args()


async def seed_kris(force: bool = False, report_file: str | None = None, excel_path: str | None = None):
    """
    Seed KRIs from Excel file.
    
    Args:
        force: If True, actually delete existing KRIs and create new ones.
               If False, run in dry-run mode (log only).
        report_file: Optional path to write unmatched KRIs for manual review.
        excel_path: Optional path to Excel file.
    """
    # Deferred imports so --help works without app modules
    import openpyxl
    from sqlalchemy import select, delete
    from app.db.session import async_session_maker
    from app.models import KeyRiskIndicator, Risk
    
    if excel_path:
        excel = Path(excel_path)
    else:
        excel = Path(__file__).parent.parent.parent / "placeholder-kri-source.xlsx"
    
    if not excel.exists():
        print(f"❌ Excel file not found: {excel}")
        return 1
    
    if not force:
        print("🔍 DRY-RUN MODE: No changes will be made. Use --force to actually seed.\n")
    
    wb = openpyxl.load_workbook(excel, data_only=True)
    unmatched_kris = []  # Track KRIs that couldn't be matched
    
    async with async_session_maker() as session:
        # First, verify risks exist
        result = await session.execute(select(Risk))
        risks = list(result.scalars().all())
        
        if not risks:
            print("❌ No risks found in database. Cannot seed KRIs.")
            print("   Please seed risks first before running this script.")
            return 1
        
        print(f"✓ Found {len(risks)} risks in database")
        
        if force:
            # Delete existing KRIs only with --force
            await session.execute(delete(KeyRiskIndicator))
            await session.commit()
            print("   ✓ Cleared existing KRIs")
        else:
            # Count existing KRIs for dry-run info
            existing_count = await session.execute(select(KeyRiskIndicator))
            existing = len(list(existing_count.scalars().all()))
            print(f"   ℹ Would delete {existing} existing KRIs (use --force to confirm)")
        
        # Group risks by category and process for better matching
        risks_by_cat = defaultdict(list)
        risks_by_proc = defaultdict(list)
        for r in risks:
            if r.category:
                risks_by_cat[r.category].append(r)
            if r.process:
                risks_by_proc[r.process].append(r)
        
        # Counters for round-robin distribution within matched categories
        cat_index = defaultdict(int)
        proc_index = defaultdict(int)
        
        created = 0
        skipped = 0
        
        for sheet_name, db_cat in SHEET_TO_CATEGORY.items():
            if sheet_name not in wb.sheetnames:
                continue
            
            print(f"\n📊 Processing: {sheet_name} -> {db_cat}")
            sheet = wb[sheet_name]
            target_risks = risks_by_cat.get(db_cat, [])
            
            if not target_risks:
                print(f"   ⚠ No risks found for category '{db_cat}' - KRIs will be logged")
            
            for row_idx in range(2, sheet.max_row + 1):
                row = list(sheet.iter_rows(min_row=row_idx, max_row=row_idx, values_only=True))[0]
                if not row or len(row) < 12:
                    continue
                
                risk_desc_raw = str(row[5]) if row[5] else ""
                metric_name = str(row[8]) if row[8] else None
                if not metric_name or metric_name.lower() in ('none', 'metrika'):
                    skipped += 1
                    continue
                
                current_value = safe_float(row[9])
                lower_limit = safe_float(row[10])
                upper_limit = safe_float(row[11], 999999)
                
                # Match strategy:
                # 1. Try to find a risk within the specific category by keyword
                # 2. Try to match by process keyword if category lookup failed
                # 3. Use round-robin within category if many risks exist
                # 4. If still no match, LOG and SKIP (no global fallback)
                
                matched_risk = None
                
                # Attempt keyword match within target category
                if target_risks:
                    words = risk_desc_raw.lower().split()
                    for word in words:
                        if len(word) > 4:
                            for r in target_risks:
                                if word in r.description.lower():
                                    matched_risk = r
                                    break
                            if matched_risk:
                                break
                    
                    # If no keyword match, use round-robin within category
                    if not matched_risk:
                        idx = cat_index[db_cat] % len(target_risks)
                        matched_risk = target_risks[idx]
                        cat_index[db_cat] += 1
                
                # If still no match, try process keywords
                if not matched_risk:
                    for kw, proc in KEYWORD_TO_PROCESS.items():
                        if kw in risk_desc_raw.lower() or kw in metric_name.lower():
                            proc_risks = risks_by_proc.get(proc)
                            if proc_risks:
                                idx = proc_index[proc] % len(proc_risks)
                                matched_risk = proc_risks[idx]
                                proc_index[proc] += 1
                                break
                
                # NO GLOBAL FALLBACK - Log unmatched and skip
                if not matched_risk:
                    unmatched_entry = {
                        "sheet": sheet_name,
                        "row": row_idx,
                        "metric_name": metric_name[:80] if metric_name else "N/A",
                        "risk_desc": risk_desc_raw[:80] if risk_desc_raw else "N/A",
                    }
                    unmatched_kris.append(unmatched_entry)
                    print(f"   ⚠ SKIPPED (no match): Row {row_idx} - '{metric_name[:40]}...'")
                    skipped += 1
                    continue
                
                unit = "%" if (current_value < 2 or "%" in metric_name) else ""
                
                if force:
                    kri = KeyRiskIndicator(
                        risk_id=matched_risk.id,
                        metric_name=metric_name[:200],
                        current_value=current_value,
                        lower_limit=lower_limit,
                        upper_limit=upper_limit,
                        unit=unit,
                    )
                    session.add(kri)
                
                created += 1
        
        if force:
            await session.commit()
    
    # Summary
    print(f"\n{'='*50}")
    if force:
        print(f"✅ Created: {created} KRIs")
    else:
        print(f"🔍 Would create: {created} KRIs (dry-run)")
    print(f"⚠  Skipped: {skipped} rows (no metric or no match)")
    print(f"❌ Unmatched: {len(unmatched_kris)} KRIs (need manual mapping)")
    
    # Write unmatched report if requested
    if report_file and unmatched_kris:
        with open(report_file, 'w') as f:
            f.write("# Unmatched KRIs - Manual Mapping Required\n\n")
            f.write("The following KRIs could not be matched to risks.\n")
            f.write("Please review and manually assign them.\n\n")
            for entry in unmatched_kris:
                f.write(f"Sheet: {entry['sheet']}, Row: {entry['row']}\n")
                f.write(f"  Metric: {entry['metric_name']}\n")
                f.write(f"  Risk Desc: {entry['risk_desc']}\n\n")
        print(f"\n📄 Unmatched KRIs written to: {report_file}")
    elif unmatched_kris and not report_file:
        print(f"\n💡 Tip: Use --report <file> to save unmatched KRIs for review")
    
    return 0


if __name__ == "__main__":
    args = parse_args()
    sys.exit(asyncio.run(seed_kris(
        force=args.force,
        report_file=args.report,
        excel_path=args.excel
    )))
