"""One-time, OUT-OF-RUNTIME cutover import of the DORA ICT Register workbook (issue #53).

Feeds the workbook's full register — Vendors (all 30 incl. the DOD stubs),
the BIZ DATA master Contract, Processes, Assets, every Link relation
(05/06/10/11 §1), Threats, and the 8 ICT Risks — through the SERVICE LAYER
(the lifecycle create/update/link functions with a real session and the
seeded risk-manager user), never through raw model writes. No upload endpoint
exists or is introduced; the no-Excel posture stands: the machine-readable
source is the workbook BUILDER's data module (``builder/seed.py`` +
``source_data.json``), not xlsx parsing, and openpyxl is never imported (a
pure column-letter stub satisfies the builder module's import).

Idempotent upsert-by-natural-key (the seed_e2e_ict_register pattern): every
row is keyed by its workbook natural key, re-runs converge (created=0), so a
failed run is simply re-run to completion — no partial state survives.
Anything the service layer rejects is a REPORTED data finding, never a
silent skip.

Usage (from backend/, DATABASE_URL required — the script refuses to guess):

    DATABASE_URL=postgresql+asyncpg://... ./venv/bin/python -m scripts.import_ict_register_workbook \
        --source "/path/to/dora-registr-aktiv-2026"

    ... --verify   # post-import fidelity characterization (no writes):
                   # asserts the workbook's documented profile — row counts,
                   # 79 CIF processes, 26 critical vendors, 358 §1 pairs,
                   # 106 §2 pairs, and all 52 DQ counts vs build_expected.json.
"""

from __future__ import annotations

import argparse
import asyncio
import importlib.util
import json
import os
import sys
import types
from dataclasses import dataclass
from datetime import date
from decimal import Decimal
from pathlib import Path
from types import ModuleType
from typing import Any

from sqlalchemy import func, select

from app.core.config import get_settings
from app.core.exceptions import AuthorizationError, ConflictError, NotFoundError, ValidationError
from app.core.user_query_options import user_selectinload_options
from app.db.session import session_context
from app.models import (
    Asset,
    AssetAssetLink,
    AssetVendorLink,
    GlobalConfig,
    Process,
    ProcessAssetLink,
    ProcessVendorLink,
    Risk,
    RiskAssetLink,
    RiskProcessLink,
    Threat,
    ThreatRiskLink,
    User,
    Vendor,
    VendorContract,
    VendorRiskLink,
    VendorSubOutsourcing,
)
from app.schemas.asset import (
    AssetCreate,
    AssetUpdate,
    AssetVendorLinkCreate,
    ProcessAssetLinkCreate,
    ProcessAssetLinkUpdate,
)
from app.schemas.process import ProcessCreate, ProcessUpdate, ProcessVendorLinkCreate
from app.schemas.risk import RiskAssetLinkCreate, RiskCreate, RiskProcessLinkCreate
from app.schemas.threat import RiskThreatLinkCreate, ThreatCreate, ThreatUpdate
from app.schemas.vendor import VendorCreate, VendorUpdate
from app.schemas.vendor_contract import VendorContractCreate, VendorContractUpdate
from app.services._entity_mutation_lifecycle.lifecycle import create_risk_detail, update_risk_detail
from app.services._ict_register_lifecycle.asset_lifecycle import create_asset_detail, update_asset_detail
from app.services._ict_register_lifecycle.asset_links import add_asset_process_link, update_asset_process_link
from app.services._ict_register_lifecycle.derivation import derive_ict_register
from app.services._ict_register_lifecycle.derivation_inputs import load_ict_register_dq_graph
from app.services._ict_register_lifecycle.dq import derive_ict_register_dq, risk_net_band, risk_vs_tolerance
from app.services._ict_register_lifecycle.lifecycle import create_process_detail, update_process_detail
from app.services._ict_register_lifecycle.risk_links import add_risk_asset_link, add_risk_process_link
from app.services._ict_register_lifecycle.threat_lifecycle import create_threat_detail, update_threat_detail
from app.services._ict_register_lifecycle.threat_links import add_risk_threat_link
from app.services._ict_register_lifecycle.vendor_links import add_asset_vendor_link, add_process_vendor_link
from app.services._ict_register_reference.parameters import (
    ICT_PARAMETER_CONFIG_CATEGORY,
    ICT_WORKBOOK_PARAMETERS_BY_NAME,
    load_ict_workbook_parameter_set,
)
from app.services._vendor_governance.contract_lifecycle import (
    create_vendor_contract_detail,
    update_vendor_contract_detail,
)
from app.services._vendor_governance.lifecycle import create_vendor_detail, update_vendor_detail
from app.services._vendor_links.workflow import link_vendor_target
from scripts._ict_register_import_helpers import (
    RiskBandScale,
    asset_preliminary_criticality,
    factor_score_for_app,
    get_column_letter,
    join_aliases,
    licensed_activity_for_l0,
    normalize_l2,
    scale_risk_band_thresholds,
    workbook_risk_scores,
    workbook_subject_value,
)

# The authorized import actor: the seeded risk manager (maintenance role per #38).
IMPORT_USER_EMAIL = "risk.manager@riskhub.local"

# App-required Vendor/Risk classification fields the workbook does not carry.
# Documented in docs/dora-ict-register/cutover-record.md.
VENDOR_PROCESS_LABEL = "ICT registr"
VENDOR_TYPE = "ict"
RISK_PROCESS_LABEL = "ICT registr"

# 05_Vazby significance is re-seeded to the closed-list bootstrap value (README v3).
VPA_SIGNIFICANCE = "Neposouzeno"
# 11 §1 imported pairs carry the workbook's review note in the Poznámka column.
PV_LINK_NOTE = "k revizi"

# The four risk-band parameters that scale at cutover (see cutover-record.md).
SCALED_PARAMETER_NAMES = ("P_RizStr", "P_RizVys", "P_RizKrit", "P_Tolerance")

DOMAIN_ERRORS = (AuthorizationError, ConflictError, NotFoundError, ValidationError)

CHECK_IDS = tuple(f"DQ-{n:02d}" for n in range(1, 53))


@dataclass
class Counters:
    created: int = 0
    updated: int = 0
    unchanged: int = 0

    def line(self) -> str:
        return f"created={self.created}, updated={self.updated}, unchanged={self.unchanged}"


class ImportReport:
    """Collects per-register counters and reported data findings."""

    def __init__(self) -> None:
        self.registers: dict[str, Counters] = {}
        self.findings: list[str] = []

    def counters(self, register: str) -> Counters:
        return self.registers.setdefault(register, Counters())

    def finding(self, message: str) -> None:
        self.findings.append(message)
        print(f"   !! FINDING: {message}")

    @property
    def total_created(self) -> int:
        return sum(c.created for c in self.registers.values())


# ---------------------------------------------------------------------------
# Builder source loading (machine-readable data module, never the xlsx).
# ---------------------------------------------------------------------------


def load_builder_seed(source_dir: Path) -> ModuleType:
    """Import the external builder's ``seed.py`` as the structured data source.

    ``seed.py`` imports ``openpyxl.utils.get_column_letter`` for its layout
    dicts (which the import never consumes). The workbook binary is never
    opened and no xlsx parsing happens: when openpyxl is absent (it is only
    a dev/test dependency here — the runtime Excel ban stands) a pure,
    characterization-tested stub satisfies the builder module's one import.
    """
    seed_path = source_dir / "builder" / "seed.py"
    if not seed_path.is_file():
        raise SystemExit(f"Builder seed module not found: {seed_path}")
    if "openpyxl.utils" not in sys.modules:
        try:
            import openpyxl.utils  # noqa: F401
        except ImportError:
            utils_stub = types.ModuleType("openpyxl.utils")
            utils_stub.get_column_letter = get_column_letter  # type: ignore[attr-defined]
            package_stub = types.ModuleType("openpyxl")
            package_stub.utils = utils_stub  # type: ignore[attr-defined]
            sys.modules["openpyxl"] = package_stub
            sys.modules["openpyxl.utils"] = utils_stub
    spec = importlib.util.spec_from_file_location("dora_workbook_builder_seed", seed_path)
    if spec is None or spec.loader is None:
        raise SystemExit(f"Cannot load builder seed module from {seed_path}")
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module


def _iso(value: str | None) -> date | None:
    return date.fromisoformat(value) if value else None


def builder_parameter_values(seed: ModuleType) -> dict[str, object]:
    """All 23 workbook parameters from the builder, typed like the app registry."""
    values: dict[str, object] = {}
    for name, value, _comment in seed.PARAMS:
        values[name] = int(value)
    for name, value, _comment in seed.PARAM_TXT:
        values[name] = str(value)
    for name, value, _comment in seed.PARAM_DATE:
        values[name] = date.fromisoformat(value)
    return values


def derive_scales(seed: ModuleType) -> tuple[RiskBandScale, RiskBandScale, int, int]:
    """Derive the workbook risk scale and its proportional app-scale counterpart.

    Workbook 13_Rizika: hrubé = hodnota_subjektu × zranitelnost × pravděpodobnost.
    Every axis tops out at 5 (Skala15; hodnota_subj maxes at MATCH over TridyKrit
    + 1 and the vendor branch's literal 5) -> scale max 5*5*5 = 125.
    App Risk: net/gross = probability × impact, both 1-5 -> scale max 25.
    """
    params = builder_parameter_values(seed)
    axis_max = max(seed.ENUMS["Skala15"])
    subject_max = max(len(seed.ENUMS["TridyKrit"]) + 1, 5)
    workbook_max = subject_max * axis_max * axis_max
    app_max = axis_max * axis_max
    workbook_scale = RiskBandScale(
        medium_from=int(params["P_RizStr"]),
        high_from=int(params["P_RizVys"]),
        critical_from=int(params["P_RizKrit"]),
        tolerance=int(params["P_Tolerance"]),
    )
    app_scale = scale_risk_band_thresholds(
        workbook_scale, workbook_scale_max=workbook_max, app_scale_max=app_max
    )
    return workbook_scale, app_scale, workbook_max, app_max


# ---------------------------------------------------------------------------
# Import actor.
# ---------------------------------------------------------------------------


async def load_import_user(db) -> User:
    result = await db.execute(
        select(User)
        .options(*user_selectinload_options(include_permissions=True))
        .where(User.email == IMPORT_USER_EMAIL)
    )
    user = result.scalar_one_or_none()
    if user is None:
        raise SystemExit(
            f"Import user {IMPORT_USER_EMAIL!r} not found — run `python -m app.db.seed` first."
        )
    return user


# ---------------------------------------------------------------------------
# Step 0 — ADR-008 parameter cutover (config overlay).
# ---------------------------------------------------------------------------


async def apply_parameter_overlay(db, seed: ModuleType, user: User, report: ImportReport) -> None:
    """Cross-check all 23 builder parameters and seed the 4 scaled overrides.

    Non-scaled parameters must equal the app registry's verbatim defaults —
    a mismatch means the engine would derive against a different methodology
    than the workbook used, so the import aborts. The four risk-band
    parameters get proportional ``global_config`` overrides (ADR-008 pattern:
    seeded row wins over the code default).
    """
    builder_values = builder_parameter_values(seed)
    mismatches = []
    for name, builder_value in builder_values.items():
        registry = ICT_WORKBOOK_PARAMETERS_BY_NAME.get(name)
        if registry is None:
            mismatches.append(f"{name}: unknown to the app parameter registry")
        elif registry.default != builder_value:
            mismatches.append(f"{name}: builder={builder_value!r} app-default={registry.default!r}")
    if mismatches:
        raise SystemExit(
            "Workbook parameters diverge from the app registry defaults — PM decision needed:\n  "
            + "\n  ".join(mismatches)
        )

    workbook_scale, app_scale, workbook_max, app_max = derive_scales(seed)
    scaled_values = {
        "P_RizStr": app_scale.medium_from,
        "P_RizVys": app_scale.high_from,
        "P_RizKrit": app_scale.critical_from,
        "P_Tolerance": app_scale.tolerance,
    }
    counters = report.counters("parameter overlay (global_config)")
    for name in SCALED_PARAMETER_NAMES:
        parameter = ICT_WORKBOOK_PARAMETERS_BY_NAME[name]
        value = str(scaled_values[name])
        description = (
            f"ICT Register cutover (#53): {name} scaled proportionally from the workbook default "
            f"{builder_values[name]} on the 1-{workbook_max} three-factor scale to the app's "
            f"1-{app_max} net-score scale. {parameter.meaning}"
        )
        existing = (
            await db.execute(select(GlobalConfig).where(GlobalConfig.key == parameter.config_key))
        ).scalar_one_or_none()
        if existing is None:
            db.add(
                GlobalConfig(
                    key=parameter.config_key,
                    value=value,
                    value_type="int",
                    category=ICT_PARAMETER_CONFIG_CATEGORY,
                    display_name=f"ICT Register parameter {name}",
                    description=description,
                    is_editable=True,
                    updated_by_id=user.id,
                )
            )
            counters.created += 1
        elif existing.value != value:
            existing.value = value
            existing.description = description
            existing.updated_by_id = user.id
            counters.updated += 1
        else:
            counters.unchanged += 1
    await db.commit()
    print(
        f"   Parameter cutover: P_RizStr {workbook_scale.medium_from}->{app_scale.medium_from}, "
        f"P_RizVys {workbook_scale.high_from}->{app_scale.high_from}, "
        f"P_RizKrit {workbook_scale.critical_from}->{app_scale.critical_from}, "
        f"P_Tolerance {workbook_scale.tolerance}->{app_scale.tolerance} "
        f"(scale {workbook_max}->{app_max}); other 19 parameters match the registry defaults."
    )


# ---------------------------------------------------------------------------
# Generic upsert plumbing.
# ---------------------------------------------------------------------------


def _differs(row: Any, target: dict[str, Any]) -> bool:
    for field, value in target.items():
        current = getattr(row, field)
        if isinstance(current, Decimal) or isinstance(value, Decimal):
            if (current is None) != (value is None):
                return True
            if current is not None and Decimal(current) != Decimal(value):
                return True
            continue
        if current != value:
            return True
    return False


# ---------------------------------------------------------------------------
# Registers.
# ---------------------------------------------------------------------------


def _vendor_rows(seed: ModuleType) -> list[tuple[str, dict[str, Any]]]:
    """All 30 vendor payloads: DOD-01 (BIZ DATA, full record) + 29 faithful stubs."""
    biz = seed.BIZ_DATA
    exante_fields = (
        "ex_ante_operational",
        "ex_ante_legal",
        "ex_ante_ict",
        "ex_ante_reputational",
        "ex_ante_data_confidentiality",
        "ex_ante_data_availability",
        "ex_ante_data_location",
        "ex_ante_provider_location",
        "ex_ante_ict_concentration",
    )
    biz_payload: dict[str, Any] = {
        "name": biz["nazev"],
        "country": biz["zeme"],
        "person_type": biz["osoba"],
        "identifier_type": biz["typk"],
        "identifier_value": biz["idk"],
        "data_storage": biz["ulozeni"],
        "service_country": biz["zposk"],
        "data_location": biz["lokd"],
        "processing_location": biz["lokz"],
        "data_sensitivity": biz["citlivost"],
        "replaceability": biz["subst"],
        "substitutability_reason": biz["duvod"],
        "last_audit_date": _iso(biz["audit"]),
        "exit_plan_state": biz["exit"],
        "reintegration": biz["reint"],
        "service_disruption_impact": biz["dopad"],
        "alternative_providers": biz["alt"],
        "ctpp_designation": biz["ctpp"],
        **dict(zip(exante_fields, biz["exante"], strict=True)),
        "ex_ante_assessment_date": _iso(biz["exante_datum"]),
        "assessment_phase": biz["faze"],
        "due_diligence_state": biz["dd"],
        "last_monitoring_date": _iso(biz["monitoring"]),
    }
    rows = [(biz["nazev"], biz_payload)]
    for provider in seed.SRC["providers"]:
        rows.append(
            (
                provider["display"],
                {
                    "name": provider["display"],
                    "reference_occurrence_count": provider["occ"],
                    "reference_process_count": provider["nproc"],
                },
            )
        )
    return rows


async def import_vendors(db, seed: ModuleType, user: User, report: ImportReport) -> dict[str, int]:
    """07_Dodavatelé — returns vendor name -> id."""
    counters = report.counters("vendors (07)")
    vendor_ids: dict[str, int] = {}
    required = {
        "process": VENDOR_PROCESS_LABEL,
        "outsourcing_owner_user_id": user.id,
        "vendor_type": VENDOR_TYPE,
    }
    for name, payload in _vendor_rows(seed):
        target = {**payload, **required}
        existing_id = (
            await db.execute(select(Vendor.id).where(Vendor.name == name))
        ).scalar_one_or_none()
        try:
            if existing_id is None:
                created = await create_vendor_detail(db=db, payload=VendorCreate(**target), current_user=user)
                vendor_ids[name] = created.id
                counters.created += 1
            else:
                vendor_ids[name] = existing_id
                row = (await db.execute(select(Vendor).where(Vendor.id == existing_id))).scalar_one()
                compare = {k: v for k, v in target.items() if k != "vendor_type"}
                if _differs(row, compare) or row.vendor_type != VENDOR_TYPE:
                    await update_vendor_detail(
                        db=db, vendor_id=existing_id, payload=VendorUpdate(**target), current_user=user
                    )
                    counters.updated += 1
                else:
                    counters.unchanged += 1
        except DOMAIN_ERRORS as error:
            report.finding(f"Vendor {name!r} rejected by the service layer: {error}")
    print(f"   Vendors: {counters.line()}")
    return vendor_ids


async def import_contracts(
    db, seed: ModuleType, user: User, vendor_ids: dict[str, int], report: ImportReport
) -> None:
    """08_Smlouvy — the single seeded BIZ DATA master contract."""
    counters = report.counters("contracts (08)")
    biz = seed.BIZ_DATA
    vendor_id = vendor_ids[biz["nazev"]]
    target: dict[str, Any] = {
        "contract_reference": biz["sml"],
        "internal_contract_number": None,
        "records_system": None,
        "arrangement_type": biz["ujedn"],
        "main_contract": "Ano",
        "overarching_arrangement_reference": None,
        "description": "Provoz a rozvoj core pojistného systému Veris",
        "roi_scope": "Ano",
        "start_date": _iso(biz["start"]),
        "end_date": _iso(biz["konec"]),
        "notice_period_entity_days": biz["vyp_e"],
        "notice_period_provider_days": biz["vyp_p"],
        "governing_law_country": biz["pravo"],
        "annual_cost": Decimal(biz["naklad"]),
        "currency": biz["mena"],
        "note": None,
    }
    existing = (
        await db.execute(
            select(VendorContract).where(
                VendorContract.vendor_id == vendor_id,
                VendorContract.contract_reference == biz["sml"],
            )
        )
    ).scalar_one_or_none()
    try:
        if existing is None:
            await create_vendor_contract_detail(
                db=db, vendor_id=vendor_id, payload=VendorContractCreate(**target), current_user=user
            )
            counters.created += 1
        elif _differs(existing, target):
            await update_vendor_contract_detail(
                db=db,
                vendor_id=vendor_id,
                contract_id=existing.id,
                payload=VendorContractUpdate(**target),
                current_user=user,
            )
            counters.updated += 1
        else:
            counters.unchanged += 1
    except DOMAIN_ERRORS as error:
        report.finding(f"Contract {biz['sml']!r} rejected by the service layer: {error}")
    print(f"   Contracts: {counters.line()} (09_Subdodávky ships empty — nothing to import)")


def _process_payload(seed: ModuleType, row: dict[str, Any]) -> dict[str, Any]:
    owner = row["owner"] or None
    return {
        "l0_area": row["l0"],
        "l1_process": row["l1"],
        "l2_subprocess": normalize_l2(row["l2"]),
        "owner": owner,
        "owner_department": seed.OWNER_UTVAR_MAP.get(owner) if owner else None,
        "impact_client": None,
        "impact_market_operations": None,
        "impact_regulatory": None,
        "impact_financial": None,
        "impact_reputational": None,
        "mtpd_hours": None,
        "preliminary_criticality": row["src_class"] or None,
        "cif_override": row["kdf_override"] or None,
        "licensed_activity": licensed_activity_for_l0(row["l0"]),
        "rto_hours": None,
        "rpo_hours": None,
        "bcm_link": row["bcm"] or None,
        "last_dr_test_date": None,
        "dr_test_result": None,
        "interruption_impact": None,
        "assessment_date": None,
        "notes": None,
    }


ProcessKey = tuple[str, str, str | None]


async def import_processes(
    db, seed: ModuleType, user: User, report: ImportReport
) -> dict[ProcessKey, int]:
    """03_Procesy — natural key (l0, l1, l2); l1 alone is NOT unique in the workbook."""
    counters = report.counters("processes (03)")
    process_ids: dict[ProcessKey, int] = {}
    for row in seed.SRC["processes"]:
        key: ProcessKey = (row["l0"], row["l1"], normalize_l2(row["l2"]))
        target = _process_payload(seed, row)
        query = select(Process).where(Process.l0_area == key[0], Process.l1_process == key[1])
        query = (
            query.where(Process.l2_subprocess.is_(None))
            if key[2] is None
            else query.where(Process.l2_subprocess == key[2])
        )
        existing = (await db.execute(query)).scalar_one_or_none()
        try:
            if existing is None:
                created = await create_process_detail(
                    db=db, payload=ProcessCreate(**target), current_user=user
                )
                process_ids[key] = created.id
                counters.created += 1
            else:
                process_ids[key] = existing.id
                if _differs(existing, target):
                    await update_process_detail(
                        db=db, process_id=existing.id, payload=ProcessUpdate(**target), current_user=user
                    )
                    counters.updated += 1
                else:
                    counters.unchanged += 1
        except DOMAIN_ERRORS as error:
            report.finding(f"Process {key!r} rejected by the service layer: {error}")
    print(f"   Processes: {counters.line()}")
    return process_ids


def _asset_payload(seed: ModuleType, row: dict[str, Any]) -> dict[str, Any]:
    is_veris = row["key"] == "veris"
    overlay = seed.VERIS_OVERLAY
    payload: dict[str, Any] = {
        "name": row["display"],
        "asset_type": overlay["typ"] if is_veris else row["typ"],
        "asset_level": None,
        "description": None,
        "physical_location": None,
        "deployment_model": None,
        "alternative_names": join_aliases(row["aliases"]),
        "business_owner": None,
        "owner_department": None,
        "ict_owner": None,
        "gdpr_relevance": row["gdpr"] or None,
        "ai_relevance": row["ai"] or None,
        "data_classification": None,
        "confidentiality_rating": None,
        "integrity_rating": None,
        "availability_rating": None,
        "authenticity_rating": None,
        "impact_client": None,
        "impact_regulatory": None,
        "substitutability_rating": None,
        "vendor_dependency_rating": None,
        "internet_exposed": None,
        "preliminary_criticality": asset_preliminary_criticality(
            row.get("bia_crit"), row["src_class"], seed.BIA_CRIT_TO_TRIDA
        ),
        "lifecycle_state": None,
        "standard_support_end_date": None,
        "extended_support_end_date": None,
        "custom_support_end_date": None,
        "last_legacy_risk_assessment_date": None,
        "review_state": "K revizi" if row["conflicts"] else None,
        "notes": None,
    }
    if is_veris:
        payload.update(
            {
                "asset_level": overlay["uroven"],
                "description": overlay["popis"],
                "physical_location": overlay["umisteni"],
                "deployment_model": overlay["nasazeni"],
                "business_owner": overlay["bus"],
                "ict_owner": overlay["ict"],
                "data_classification": overlay["klasdat"],
                "confidentiality_rating": overlay["C"],
                "integrity_rating": overlay["I"],
                "availability_rating": overlay["A"],
                "authenticity_rating": overlay["au"],
                "impact_client": overlay["klient"],
                "impact_regulatory": overlay["reg"],
                "substitutability_rating": overlay["nahr"],
                "vendor_dependency_rating": overlay["zavis"],
                "internet_exposed": overlay["internet"],
                "lifecycle_state": overlay["stav"],
            }
        )
    else:
        owner = row["owner"] or None
        payload.update(
            {
                "business_owner": owner,
                "owner_department": seed.OWNER_UTVAR_MAP.get(owner) if owner else None,
                "lifecycle_state": "V provozu",
            }
        )
    return payload


async def import_assets(db, seed: ModuleType, user: User, report: ImportReport) -> dict[str, int]:
    """04_Aktiva — natural key = display name (unique); returns builder asset KEY -> id."""
    counters = report.counters("assets (04)")
    asset_ids: dict[str, int] = {}
    for row in seed.SRC["assets"]:
        target = _asset_payload(seed, row)
        existing = (
            await db.execute(select(Asset).where(Asset.name == row["display"]))
        ).scalar_one_or_none()
        try:
            if existing is None:
                created = await create_asset_detail(db=db, payload=AssetCreate(**target), current_user=user)
                asset_ids[row["key"]] = created.id
                counters.created += 1
            else:
                asset_ids[row["key"]] = existing.id
                if _differs(existing, target):
                    await update_asset_detail(
                        db=db, asset_id=existing.id, payload=AssetUpdate(**target), current_user=user
                    )
                    counters.updated += 1
                else:
                    counters.unchanged += 1
        except DOMAIN_ERRORS as error:
            report.finding(f"Asset {row['display']!r} rejected by the service layer: {error}")
    print(f"   Assets: {counters.line()}")
    return asset_ids


async def import_process_asset_links(
    db,
    seed: ModuleType,
    user: User,
    process_ids: dict[ProcessKey, int],
    asset_ids: dict[str, int],
    report: ImportReport,
) -> None:
    """05_Vazby_proces_aktivum — 1000 links, significance 'Neposouzeno'.

    The asset's primary-Process designation (04!proc_id, the build-time pick
    promoted to a user-controlled attribute per #38) rides ``is_primary`` on
    the matching link; DQ-07 guarantees the primary pair exists in 05.
    """
    counters = report.counters("process-asset links (05)")
    primary_pairs: set[tuple[ProcessKey, str]] = set()
    for asset in seed.SRC["assets"]:
        l0, l1, l2 = asset["primary"]
        primary_pairs.add(((l0, l1, normalize_l2(l2)), asset["key"]))

    seen_pairs: set[tuple[ProcessKey, str]] = set()
    for link in seed.SRC["vpa"]:
        key: ProcessKey = (link["l0"], link["l1"], normalize_l2(link["l2"]))
        pair = (key, link["akt"])
        seen_pairs.add(pair)
        process_id = process_ids.get(key)
        asset_id = asset_ids.get(link["akt"])
        if process_id is None or asset_id is None:
            report.finding(f"05 link {pair!r} references a row that failed to import")
            continue
        target = {
            "significance": VPA_SIGNIFICANCE,
            "spof": None,
            "is_primary": pair in primary_pairs,
            "note": None,
        }
        existing = (
            await db.execute(
                select(ProcessAssetLink).where(
                    ProcessAssetLink.process_id == process_id,
                    ProcessAssetLink.asset_id == asset_id,
                )
            )
        ).scalar_one_or_none()
        try:
            if existing is None:
                await add_asset_process_link(
                    db,
                    asset_id=asset_id,
                    payload=ProcessAssetLinkCreate(process_id=process_id, **target),
                    current_user=user,
                )
                counters.created += 1
            elif _differs(existing, target):
                await update_asset_process_link(
                    db,
                    asset_id=asset_id,
                    process_id=process_id,
                    payload=ProcessAssetLinkUpdate(**target),
                    current_user=user,
                )
                counters.updated += 1
            else:
                counters.unchanged += 1
        except DOMAIN_ERRORS as error:
            report.finding(f"05 link {pair!r} rejected by the service layer: {error}")

    missing_primaries = primary_pairs - seen_pairs
    for pair in sorted(missing_primaries):
        report.finding(f"Asset {pair[1]!r} primary Process {pair[0]!r} has no 05 link (DQ-07 material)")
    print(f"   Process-Asset links: {counters.line()} (06_Vazby_aktivum_aktivum ships empty)")


async def import_asset_vendor_links(
    db,
    seed: ModuleType,
    user: User,
    vendor_ids: dict[str, int],
    asset_ids: dict[str, int],
    report: ImportReport,
) -> None:
    """10_Vazby_aktivum_dodavatel — the two seeded Veris -> BIZ DATA service links."""
    counters = report.counters("asset-vendor links (10)")
    biz_name = seed.BIZ_DATA["nazev"]
    # Builder sheets_vendors.py build_vad seeds exactly these two rows.
    seeded = [
        ("veris", biz_name, "Dodává", "S02", "SML-2020-001", "Úplná závislost"),
        ("veris", biz_name, "Spravuje", "S14", "SML-2020-001", "Úplná závislost"),
    ]
    for asset_key, vendor_name, role, s_code, contract_ref, reliance in seeded:
        asset_id = asset_ids[asset_key]
        vendor_id = vendor_ids[vendor_name]
        existing = (
            await db.execute(
                select(AssetVendorLink).where(
                    AssetVendorLink.asset_id == asset_id,
                    AssetVendorLink.vendor_id == vendor_id,
                    AssetVendorLink.ict_service_code == s_code,
                )
            )
        ).scalar_one_or_none()
        target = {"vendor_role": role, "contract_reference": contract_ref, "reliance": reliance, "note": None}
        try:
            if existing is None:
                await add_asset_vendor_link(
                    db,
                    asset_id=asset_id,
                    payload=AssetVendorLinkCreate(vendor_id=vendor_id, ict_service_code=s_code, **target),
                    current_user=user,
                )
                counters.created += 1
            elif _differs(existing, target):
                report.finding(
                    f"10 link ({asset_key}, {vendor_name}, {s_code}) drifted from the workbook values "
                    "— no update surface exists for Asset-Vendor links; reconcile manually"
                )
            else:
                counters.unchanged += 1
        except DOMAIN_ERRORS as error:
            report.finding(f"10 link ({asset_key}, {s_code}) rejected by the service layer: {error}")
    print(f"   Asset-Vendor links: {counters.line()}")


async def import_process_vendor_links(
    db,
    seed: ModuleType,
    user: User,
    process_ids: dict[ProcessKey, int],
    vendor_ids: dict[str, int],
    report: ImportReport,
) -> None:
    """11 §1 — 358 direct Process<->Vendor pairs, note 'k revizi' (workbook column H)."""
    counters = report.counters("process-vendor §1 links (11)")
    provider_display = {p["key"]: p["display"] for p in seed.SRC["providers"]}
    for pair in seed.SRC["vpd_direct"]:
        key: ProcessKey = (pair["l0"], pair["l1"], normalize_l2(pair["l2"]))
        process_id = process_ids.get(key)
        vendor_id = vendor_ids.get(provider_display[pair["prov"]])
        if process_id is None or vendor_id is None:
            report.finding(f"11 §1 pair {key!r} -> {pair['prov']!r} references a failed row")
            continue
        existing = (
            await db.execute(
                select(ProcessVendorLink).where(
                    ProcessVendorLink.process_id == process_id,
                    ProcessVendorLink.vendor_id == vendor_id,
                )
            )
        ).scalar_one_or_none()
        target = {"direct_service_description": None, "note": PV_LINK_NOTE}
        try:
            if existing is None:
                await add_process_vendor_link(
                    db,
                    process_id=process_id,
                    payload=ProcessVendorLinkCreate(vendor_id=vendor_id, **target),
                    current_user=user,
                )
                counters.created += 1
            elif _differs(existing, target):
                report.finding(
                    f"11 §1 pair {key!r} -> {pair['prov']!r} drifted from the workbook values "
                    "— no update surface exists for Process-Vendor links; reconcile manually"
                )
            else:
                counters.unchanged += 1
        except DOMAIN_ERRORS as error:
            report.finding(f"11 §1 pair {key!r} rejected by the service layer: {error}")
    print(f"   Process-Vendor §1 links: {counters.line()}")


async def import_threats(db, seed: ModuleType, user: User, report: ImportReport) -> dict[int, int]:
    """12_Hrozby — 16 curated catalog entries; returns 1-based HR index -> id."""
    counters = report.counters("threats (12)")
    threat_ids: dict[int, int] = {}
    for index, (name, category, description, weaknesses, subject) in enumerate(seed.THREATS, start=1):
        target = {
            "name": name,
            "category": category,
            "description": description,
            "typical_weaknesses": weaknesses,
            "relevant_subject": subject,
            "notes": None,
        }
        existing = (await db.execute(select(Threat).where(Threat.name == name))).scalar_one_or_none()
        try:
            if existing is None:
                created = await create_threat_detail(db=db, payload=ThreatCreate(**target), current_user=user)
                threat_ids[index] = created.id
                counters.created += 1
            else:
                threat_ids[index] = existing.id
                if _differs(existing, target):
                    await update_threat_detail(
                        db=db, threat_id=existing.id, payload=ThreatUpdate(**target), current_user=user
                    )
                    counters.updated += 1
                else:
                    counters.unchanged += 1
        except DOMAIN_ERRORS as error:
            report.finding(f"Threat {name!r} rejected by the service layer: {error}")
    print(f"   Threats: {counters.line()}")
    return threat_ids


# ---------------------------------------------------------------------------
# 13_Rizika.
# ---------------------------------------------------------------------------


@dataclass(frozen=True)
class WorkbookRisk:
    """One parsed seed.RISKS tuple (builder field order, sheets_vendors.py:715-718)."""

    code: str
    subject_type: str
    subject_key: Any
    threat_index: int
    vulnerability: int
    probability: int
    controls: str
    effectiveness: float
    response: str
    approver: str
    justification: str
    acceptance_date: str
    review_trigger: str
    phase: str
    effectiveness_check_date: str
    effectiveness_check_result: str
    material_equity: str
    material_outage: str
    assessment_date: str
    owner: str
    action_plan_date: str
    status: str


def parse_workbook_risks(seed: ModuleType) -> list[WorkbookRisk]:
    risks = []
    for index, row in enumerate(seed.RISKS, start=1):
        (
            subject_type,
            subject_key,
            threat_index,
            vulnerability,
            probability,
            controls,
            effectiveness,
            response,
            approver,
            justification,
            acceptance,
            _pz1,
            _pz2,
            _pz3,
            trigger,
            phase,
            check_date,
            check_result,
            material_equity,
            material_outage,
            assessed,
            owner,
            action_plan,
            status,
        ) = row
        risks.append(
            WorkbookRisk(
                code=f"RIZ-{index:03d}",
                subject_type=subject_type,
                subject_key=subject_key,
                threat_index=threat_index,
                vulnerability=vulnerability,
                probability=probability,
                controls=controls,
                effectiveness=effectiveness,
                response=response,
                approver=approver,
                justification=justification,
                acceptance_date=acceptance,
                review_trigger=trigger,
                phase=phase,
                effectiveness_check_date=check_date,
                effectiveness_check_result=check_result,
                material_equity=material_equity,
                material_outage=material_outage,
                assessment_date=assessed,
                owner=owner,
                action_plan_date=action_plan,
                status=status,
            )
        )
    return risks


def _risk_description(
    risk: WorkbookRisk,
    threat_name: str,
    subject_display: str,
    gross: int,
    net: int,
    workbook_scale: RiskBandScale,
    workbook_max: int,
) -> str:
    """Faithful narrative capture of every workbook 13_Rizika column without an app analog."""
    gross_band = risk_net_band(
        gross,
        medium_from=workbook_scale.medium_from,
        high_from=workbook_scale.high_from,
        critical_from=workbook_scale.critical_from,
    )
    net_band = risk_net_band(
        net,
        medium_from=workbook_scale.medium_from,
        high_from=workbook_scale.high_from,
        critical_from=workbook_scale.critical_from,
    )
    verdict = risk_vs_tolerance(net, tolerance=workbook_scale.tolerance)
    effectiveness_pct = f"{round(risk.effectiveness * 100)} %"
    lines = [
        f"Import z DORA sešitu 13_Rizika ({risk.code}).",
        f"Hrozba: {threat_name}. Subjekt: {risk.subject_type} {subject_display}.",
        f"Kontroly: {risk.controls} (účinnost {effectiveness_pct}).",
        f"Odezva: {risk.response}. Trigger přezkumu: {risk.review_trigger}. Fáze: {risk.phase}.",
        (
            f"Poslední kontrola účinnosti: {risk.effectiveness_check_date}"
            f" ({risk.effectiveness_check_result})."
        ),
        (
            f"Materialita — dopad do VK: {risk.material_equity};"
            f" výpadek > limit: {risk.material_outage}."
        ),
        (
            f"Datum posouzení: {risk.assessment_date}. Vlastník (sešit): {risk.owner}."
            f" Termín akčního plánu: {risk.action_plan_date or '—'}. Stav (sešit): {risk.status}."
        ),
        (
            f"Původní skóre sešitu (škála 1–{workbook_max}): hrubé {gross} (pásmo {gross_band}),"
            f" čisté {net} (pásmo {net_band}, {verdict})."
        ),
    ]
    return "\n".join(lines)


async def import_risks(
    db,
    seed: ModuleType,
    user: User,
    process_ids: dict[ProcessKey, int],
    asset_ids: dict[str, int],
    vendor_ids: dict[str, int],
    threat_ids: dict[int, int],
    report: ImportReport,
) -> None:
    """13_Rizika — 8 risks scored from the LIVE derived register, workbook-verbatim.

    The subject value (hodnota_subjektu) reads the subject's CURRENT derived
    class/tier from the app's own engine — exactly how the workbook's XLOOKUP
    reads its live derived columns — then hrubé/čisté follow the workbook
    formulas on the 1-125 scale and factor onto the app's two-factor 1-25
    scale with band and tolerance verdicts preserved (#50 mapping: net=ciste).
    """
    counters = report.counters("risks (13)")
    link_counters = report.counters("risk links (13)")

    workbook_scale, app_scale, workbook_max, app_max = derive_scales(seed)
    dq_graph = await load_ict_register_dq_graph(db)
    parameters = await load_ict_workbook_parameter_set(db)
    derivation = derive_ict_register(dq_graph.graph, parameters)

    asset_display = {a["key"]: a["display"] for a in seed.SRC["assets"]}
    threat_names = {index: row[0] for index, row in enumerate(seed.THREATS, start=1)}

    for risk in parse_workbook_risks(seed):
        # Resolve the subject and its live derived label.
        if risk.subject_type == "Aktivum":
            subject_id = asset_ids[risk.subject_key]
            subject_display = asset_display[risk.subject_key]
            derived_label = derivation.assets[subject_id].resulting_criticality
        elif risk.subject_type == "Dodavatel":
            subject_display = seed.BIZ_DATA["nazev"]
            subject_id = vendor_ids[subject_display]
            derived_label = derivation.vendors[subject_id].tier
        else:
            key: ProcessKey = (
                risk.subject_key[0],
                risk.subject_key[1],
                normalize_l2(risk.subject_key[2]),
            )
            subject_id = process_ids[key]
            subject_display = risk.subject_key[1]
            derived_label = derivation.processes[subject_id].criticality_class

        subject_value = workbook_subject_value(risk.subject_type, derived_label)
        if subject_value is None:
            raise SystemExit(
                f"{risk.code}: subject {risk.subject_type} {subject_display!r} has a blank derived "
                f"class — the workbook would score this risk blank, which the app cannot represent. "
                "PM decision needed."
            )
        gross, net = workbook_risk_scores(
            subject_value, risk.vulnerability, risk.probability, risk.effectiveness
        )
        gross_probability, gross_impact = factor_score_for_app(
            gross,
            workbook_scale=workbook_scale,
            app_scale=app_scale,
            workbook_scale_max=workbook_max,
            app_scale_max=app_max,
            preferred_probability=risk.probability,
            enforce_tolerance=False,
        )
        net_probability, net_impact = factor_score_for_app(
            net,
            workbook_scale=workbook_scale,
            app_scale=app_scale,
            workbook_scale_max=workbook_max,
            app_scale_max=app_max,
            preferred_probability=risk.probability,
            enforce_tolerance=True,
        )

        threat_name = threat_names[risk.threat_index]
        target: dict[str, Any] = {
            "risk_id_code": risk.code,
            "name": f"{threat_name} — {subject_display}",
            "process": RISK_PROCESS_LABEL,
            "subprocess": f"{risk.subject_type}: {subject_display}",
            "risk_type": "operational",
            "category": None,
            "description": _risk_description(
                risk, threat_name, subject_display, gross, net, workbook_scale, workbook_max
            ),
            "department_id": None,
            "owner_id": None,
            "gross_probability": gross_probability,
            "gross_impact": gross_impact,
            "net_probability": net_probability,
            "net_impact": net_impact,
            "acceptance_approver": risk.approver or None,
            "acceptance_justification": risk.justification or None,
            "acceptance_date": _iso(risk.acceptance_date),
        }
        existing = (
            await db.execute(select(Risk).where(Risk.risk_id_code == risk.code))
        ).scalar_one_or_none()
        risk_id: int | None = None
        try:
            if existing is None:
                created = await create_risk_detail(db=db, risk_data=RiskCreate(**target), current_user=user)
                risk_id = created.id
                counters.created += 1
            else:
                risk_id = existing.id
                changed = {
                    field: value for field, value in target.items() if getattr(existing, field) != value
                }
                if changed:
                    await update_risk_detail(
                        db=db, risk_id=existing.id, update_data=changed, current_user=user
                    )
                    counters.updated += 1
                else:
                    counters.unchanged += 1
        except DOMAIN_ERRORS as error:
            report.finding(f"Risk {risk.code} rejected by the service layer: {error}")
            continue

        # Link relations: Threat<->Risk plus the subject link (13_Rizika graph).
        try:
            threat_id = threat_ids[risk.threat_index]
            existing_link = (
                await db.execute(
                    select(ThreatRiskLink).where(
                        ThreatRiskLink.threat_id == threat_id, ThreatRiskLink.risk_id == risk_id
                    )
                )
            ).scalar_one_or_none()
            if existing_link is None:
                await add_risk_threat_link(
                    db, risk_id=risk_id, payload=RiskThreatLinkCreate(threat_id=threat_id), current_user=user
                )
                link_counters.created += 1
            else:
                link_counters.unchanged += 1

            if risk.subject_type == "Aktivum":
                pair = (
                    await db.execute(
                        select(RiskAssetLink).where(
                            RiskAssetLink.risk_id == risk_id, RiskAssetLink.asset_id == subject_id
                        )
                    )
                ).scalar_one_or_none()
                if pair is None:
                    await add_risk_asset_link(
                        db,
                        risk_id=risk_id,
                        payload=RiskAssetLinkCreate(asset_id=subject_id),
                        current_user=user,
                    )
                    link_counters.created += 1
                else:
                    link_counters.unchanged += 1
            elif risk.subject_type == "Proces":
                pair = (
                    await db.execute(
                        select(RiskProcessLink).where(
                            RiskProcessLink.risk_id == risk_id, RiskProcessLink.process_id == subject_id
                        )
                    )
                ).scalar_one_or_none()
                if pair is None:
                    await add_risk_process_link(
                        db,
                        risk_id=risk_id,
                        payload=RiskProcessLinkCreate(process_id=subject_id),
                        current_user=user,
                    )
                    link_counters.created += 1
                else:
                    link_counters.unchanged += 1
            else:
                pair = (
                    await db.execute(
                        select(VendorRiskLink).where(
                            VendorRiskLink.vendor_id == subject_id, VendorRiskLink.risk_id == risk_id
                        )
                    )
                ).scalar_one_or_none()
                if pair is None:
                    await link_vendor_target(
                        db, vendor_id=subject_id, current_user=user, kind="risk", entity_id=risk_id
                    )
                    link_counters.created += 1
                else:
                    link_counters.unchanged += 1
        except DOMAIN_ERRORS as error:
            report.finding(f"Risk {risk.code} link rejected by the service layer: {error}")

        print(
            f"   {risk.code}: subject={risk.subject_type} {subject_display!r} class={derived_label!r} "
            f"E={subject_value} -> workbook gross={gross}/net={net} "
            f"-> app gross={gross_probability}x{gross_impact}={gross_probability * gross_impact} "
            f"net={net_probability}x{net_impact}={net_probability * net_impact}"
        )
    print(f"   Risks: {counters.line()}; risk links: {link_counters.line()}")


# ---------------------------------------------------------------------------
# Modes.
# ---------------------------------------------------------------------------


async def run_import(source_dir: Path) -> int:
    seed = load_builder_seed(source_dir)
    report = ImportReport()
    print("=" * 78)
    print("ICT REGISTER CUTOVER IMPORT — workbook builder data through the service layer")
    print(f"Source: {source_dir}")
    print("=" * 78)
    async with session_context(get_settings()) as db:
        user = await load_import_user(db)
        print(f"Import actor: {user.email} (id={user.id})")

        await apply_parameter_overlay(db, seed, user, report)
        vendor_ids = await import_vendors(db, seed, user, report)
        await import_contracts(db, seed, user, vendor_ids, report)
        process_ids = await import_processes(db, seed, user, report)
        asset_ids = await import_assets(db, seed, user, report)
        await import_process_asset_links(db, seed, user, process_ids, asset_ids, report)
        await import_asset_vendor_links(db, seed, user, vendor_ids, asset_ids, report)
        await import_process_vendor_links(db, seed, user, process_ids, vendor_ids, report)
        threat_ids = await import_threats(db, seed, user, report)
        await import_risks(db, seed, user, process_ids, asset_ids, vendor_ids, threat_ids, report)

    print("-" * 78)
    print("IMPORT SUMMARY")
    for register, counters in report.registers.items():
        print(f"   {register}: {counters.line()}")
    print(f"   TOTAL created: {report.total_created}")
    if report.findings:
        print(f"REPORTED DATA FINDINGS ({len(report.findings)}):")
        for finding in report.findings:
            print(f"   - {finding}")
        return 2
    print("No data findings — every workbook row passed the service layer.")
    return 0


async def run_verify(source_dir: Path, expected_path: Path) -> int:
    """Post-import fidelity characterization: the workbook's documented profile, engine-asserted."""
    seed = load_builder_seed(source_dir)
    expected = json.loads(expected_path.read_text(encoding="utf-8"))
    mismatches: list[str] = []

    def check(label: str, actual: object, wanted: object) -> None:
        status = "OK " if actual == wanted else "MISMATCH"
        if actual != wanted:
            mismatches.append(f"{label}: workbook-expected {wanted!r}, app-actual {actual!r}")
        print(f"   [{status}] {label}: expected {wanted!r}, actual {actual!r}")

    print("=" * 78)
    print("ICT REGISTER FIDELITY CHARACTERIZATION (--verify, read-only)")
    print(f"Source: {source_dir}; expected profile: {expected_path}")
    print("=" * 78)
    async with session_context(get_settings()) as db:
        parameters = await load_ict_workbook_parameter_set(db)
        print(
            "Effective risk-band parameters: "
            f"P_RizStr={parameters.value('P_RizStr')}, P_RizVys={parameters.value('P_RizVys')}, "
            f"P_RizKrit={parameters.value('P_RizKrit')}, P_Tolerance={parameters.value('P_Tolerance')} "
            f"(methodology version {parameters.version})"
        )
        dq_graph = await load_ict_register_dq_graph(db)
        graph = dq_graph.graph
        derivation = derive_ict_register(graph, parameters)
        dq_result = derive_ict_register_dq(dq_graph, parameters)
        threat_count = (await db.execute(select(func.count(Threat.id)))).scalar_one()
        sub_count = (await db.execute(select(func.count(VendorSubOutsourcing.id)))).scalar_one()
        aa_count = (await db.execute(select(func.count(AssetAssetLink.id)))).scalar_one()

        profile = seed.SRC["profile"]
        print("-- Register row counts --")
        check("processes (03)", len(graph.processes), profile["processes"])
        check("assets (04)", len(graph.assets), profile["assets"])
        check("vendors (07)", len(graph.vendors), 1 + profile["providers"])
        check("contracts (08)", len(graph.contracts), 1)
        check("sub-outsourcing (09)", sub_count, 0)
        check("process-asset links (05)", len(graph.process_asset_links), profile["vpa_rows"])
        check("asset-asset links (06)", aa_count, 0)
        check("asset-vendor links (10)", len(graph.asset_vendor_links), 2)
        check("process-vendor §1 links (11)", len(graph.process_vendor_links), profile["vpd_direct"])
        check("threats (12)", threat_count, len(seed.THREATS))
        check("ICT-linked risks (13)", len(dq_graph.risks), expected["n_risks"])

        print("-- Engine-derived profile --")
        cif_processes = sum(1 for d in derivation.processes.values() if d.cif == "Ano")
        check("CIF processes (n_kdf)", cif_processes, expected["n_kdf"])
        critical_vendor_ids = sorted(
            vid for vid, d in derivation.vendors.items() if d.tier == "Kritický dodavatel"
        )
        check("critical vendors (n_krit_vendors)", len(critical_vendor_ids), expected["n_krit_vendors"])
        pairs_total = sum(len(d.transitive_process_links) for d in derivation.vendors.values())
        check("derived §2 pairs (pairs_total)", pairs_total, expected["pairs_total"])

        # Critical-vendor identity: the builder's DOD ids -> names -> app rows.
        vendor_id_by_name = {vendor.name: vendor.id for vendor in graph.vendors}
        expected_names = {seed.BIZ_DATA["nazev"]}
        provider_by_dod = {seed.dod_id_for_provider(p["key"]): p["display"] for p in seed.SRC["providers"]}
        for dod_id in expected["krit_candidates"]:
            expected_names.add(provider_by_dod[dod_id])
        expected_critical_ids = sorted(
            vendor_id_by_name[name] for name in expected_names if name in vendor_id_by_name
        )
        check("critical vendor identity (DOD-01 + krit_candidates)", critical_vendor_ids, expected_critical_ids)

        # Key derived spot values — the builder verify-gate's assertions, in-app.
        biz_vendor_id = vendor_id_by_name.get(seed.BIZ_DATA["nazev"])
        veris_asset_id = next(
            (asset.id for asset in graph.assets if asset.name == "Veris"), None
        )
        if veris_asset_id is not None:
            check("Veris resulting criticality", derivation.assets[veris_asset_id].resulting_criticality, "Kritická")
            check("Veris CIF", derivation.assets[veris_asset_id].cif, "Ano")
        if biz_vendor_id is not None:
            biz = derivation.vendors[biz_vendor_id]
            check("BIZ DATA tier", biz.tier, "Kritický dodavatel")
            check("BIZ DATA main contract propagated", biz.main_contract_reference, seed.BIZ_DATA["sml"])
            check("BIZ DATA country category", biz.country_category, "ČR")

        print("-- ICT-linked risk bands (informative) --")
        for risk in dq_graph.risks:
            band = risk_net_band(
                risk.net_score,
                medium_from=int(parameters.value("P_RizStr")),
                high_from=int(parameters.value("P_RizVys")),
                critical_from=int(parameters.value("P_RizKrit")),
            )
            verdict = risk_vs_tolerance(risk.net_score, tolerance=int(parameters.value("P_Tolerance")))
            print(f"   {risk.label}: net={risk.net_score} band={band} {verdict}")

        print("-- DQ profile: all 52 checks vs build_expected.json --")
        actual_by_id = {c.check_id: c for c in dq_result.checks}
        for check_id in CHECK_IDS:
            result = actual_by_id[check_id]
            check(f"{check_id} ({result.title_cs})", result.count, expected["dq"][check_id])
        expected_findings = sum(1 for check_id in CHECK_IDS if expected["dq"][check_id] > 0)
        check("checks with findings (non-zero)", dq_result.finding_count, expected_findings)

    print("-" * 78)
    if mismatches:
        print(f"CHARACTERIZATION MISMATCHES ({len(mismatches)}):")
        for mismatch in mismatches:
            print(f"   - {mismatch}")
        return 1
    print("CHARACTERIZATION PASSED — the register reproduces the workbook's documented profile.")
    return 0


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument(
        "--source",
        required=True,
        type=Path,
        help="External workbook export directory (contains builder/seed.py; read-only, never committed)",
    )
    parser.add_argument(
        "--verify",
        action="store_true",
        help="Run the read-only fidelity characterization instead of importing",
    )
    parser.add_argument(
        "--expected",
        type=Path,
        default=None,
        help="Expected-profile JSON (default: <source>/builder/build_expected.json)",
    )
    args = parser.parse_args()

    if not os.environ.get("DATABASE_URL"):
        raise SystemExit(
            "DATABASE_URL must be set explicitly — the cutover import refuses to guess a database."
        )
    source_dir = args.source.expanduser().resolve()
    if not source_dir.is_dir():
        raise SystemExit(f"Source directory not found: {source_dir}")

    if args.verify:
        expected_path = args.expected or (source_dir / "builder" / "build_expected.json")
        if not expected_path.is_file():
            raise SystemExit(f"Expected-profile JSON not found: {expected_path}")
        exit_code = asyncio.run(run_verify(source_dir, expected_path))
    else:
        exit_code = asyncio.run(run_import(source_dir))
    raise SystemExit(exit_code)


if __name__ == "__main__":
    main()
