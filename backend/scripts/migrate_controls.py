#!/usr/bin/env python3
"""Deterministic control import with dry-run/report/apply/reset contracts."""

from __future__ import annotations

import argparse
import asyncio
import sys
from dataclasses import dataclass
from difflib import SequenceMatcher
from pathlib import Path

import openpyxl
from sqlalchemy import delete, select

from app.core.config import get_settings
from app.db.session import session_context
from app.models import Control, ControlExecution, ControlRiskLink, Department, Risk, User
from scripts.import_contracts import ImportReport, write_report

FREQ_MAP = {
    "denne": "daily",
    "denně": "daily",
    "tydne": "weekly",
    "týdně": "weekly",
    "mesicne": "monthly",
    "měsíčně": "monthly",
    "ctvrtletne": "quarterly",
    "čtvrtletně": "quarterly",
    "rocne": "annually",
    "ročně": "annually",
    "ad hoc": "ad_hoc",
    "adhoc": "ad_hoc",
}
FORM_MAP = {
    "manualni": "manual",
    "manuální": "manual",
    "manual": "manual",
    "automatic": "automatic",
    "automaticka": "automatic",
    "automatická": "automatic",
}


@dataclass(slots=True)
class ControlImportRow:
    row_number: int
    name: str
    description: str
    data_source: str | None
    methodology_reference: str | None
    control_form: str
    process_owner_position: str | None
    executor_position: str | None
    frequency: str
    output_description: str | None
    report_recipient: str | None
    documentation_location: str | None
    risks_text: str


def _parse_args(argv: list[str]) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Import controls from an Excel workbook.")
    parser.add_argument("--input", required=True, help="Path to the source workbook.")
    parser.add_argument("--apply", action="store_true", help="Persist changes. Default is dry-run.")
    parser.add_argument(
        "--allow-reset",
        action="store_true",
        help="Allow destructive reset of existing controls, executions, and control-risk links before import.",
    )
    parser.add_argument("--report", help="Optional path for a JSON import report.")
    return parser.parse_args(argv)


def _normalize_string(value: object | None) -> str:
    if value is None:
        return ""
    return " ".join(str(value).lower().strip().replace("\n", " ").split())


def _similarity_score(left: str, right: str) -> float:
    return SequenceMatcher(None, _normalize_string(left), _normalize_string(right)).ratio()


def _optional_text(value: object | None) -> str | None:
    text = str(value).strip() if value is not None else ""
    return text or None


def _load_rows(input_path: Path, report: ImportReport) -> list[ControlImportRow]:
    workbook = openpyxl.load_workbook(input_path, data_only=True)
    sheet = workbook.active
    rows: list[ControlImportRow] = []
    seen_names: dict[str, int] = {}

    for row_idx in range(2, sheet.max_row + 1):
        raw_name = sheet.cell(row=row_idx, column=2).value
        if not raw_name:
            report.skips += 1
            continue

        name = str(raw_name).strip()
        normalized_name = _normalize_string(name)
        if normalized_name in seen_names:
            report.add_error(
                "duplicate-control-name",
                (
                    f"Workbook contains duplicate control name {name!r} on rows "
                    f"{seen_names[normalized_name]} and {row_idx}."
                ),
                row=row_idx,
            )
            continue
        seen_names[normalized_name] = row_idx

        control_form = FORM_MAP.get(
            _normalize_string(sheet.cell(row=row_idx, column=6).value or "manual"),
            "manual",
        )
        frequency = FREQ_MAP.get(
            _normalize_string(sheet.cell(row=row_idx, column=10).value or "mesicne"),
            "monthly",
        )

        description = str(sheet.cell(row=row_idx, column=3).value or "").strip()
        corrective = str(sheet.cell(row=row_idx, column=15).value or "").strip()
        if corrective:
            description = f"{description}\n\nNápravné opatření: {corrective}".strip()

        rows.append(
            ControlImportRow(
                row_number=row_idx,
                name=name,
                description=description,
                data_source=_optional_text(sheet.cell(row=row_idx, column=4).value),
                methodology_reference=_optional_text(sheet.cell(row=row_idx, column=5).value),
                control_form=control_form,
                process_owner_position=_optional_text(sheet.cell(row=row_idx, column=7).value),
                executor_position=_optional_text(sheet.cell(row=row_idx, column=9).value),
                frequency=frequency,
                output_description=_optional_text(sheet.cell(row=row_idx, column=11).value),
                report_recipient=_optional_text(sheet.cell(row=row_idx, column=12).value),
                documentation_location=_optional_text(sheet.cell(row=row_idx, column=13).value),
                risks_text=str(sheet.cell(row=row_idx, column=14).value or "").strip(),
            )
        )

    report.metadata["rows_loaded"] = len(rows)
    if not rows and report.ok:
        report.add_error("no-rows", "Workbook did not contain any importable control rows.")
    return rows


def _match_risks(risks_text: str, all_risks: list[Risk]) -> list[Risk]:
    if not risks_text:
        return []

    scored: list[tuple[Risk, float]] = []
    for risk in all_risks:
        score = _similarity_score(risks_text, risk.description or "")
        if score > 0.3:
            scored.append((risk, score))

    scored.sort(key=lambda item: (-item[1], item[0].risk_id_code))
    matched: list[Risk] = []
    seen_ids: set[int] = set()
    for risk, _score in scored:
        if risk.id in seen_ids:
            continue
        matched.append(risk)
        seen_ids.add(risk.id)
        if len(matched) == 3:
            break
    return matched


async def _run(args: argparse.Namespace) -> int:
    input_path = Path(args.input).expanduser().resolve()
    report = ImportReport(
        script="migrate_controls.py",
        input_path=str(input_path),
        apply=args.apply,
        allow_reset=args.allow_reset,
    )

    if not input_path.exists():
        report.add_error("missing-input", f"Input workbook not found: {input_path}")
        write_report(args.report, report)
        print(f"❌ Input workbook not found: {input_path}")
        return 1

    rows = _load_rows(input_path, report)

    async with session_context(get_settings()) as session:
        ops_department = (
            await session.execute(select(Department).where(Department.code == "OPS"))
        ).scalar_one_or_none()
        if ops_department is None:
            report.add_error(
                "missing-ops-department",
                "OPS department not found. Seed departments before importing controls.",
            )

        owner = (await session.execute(select(User).order_by(User.id).limit(1))).scalar_one_or_none()
        if owner is None:
            report.add_error("missing-owner", "No users found. Seed base data before importing controls.")

        all_risks = list((await session.execute(select(Risk))).scalars().all())
        if not all_risks:
            report.add_error("missing-risks", "No risks found. Import or seed risks before importing controls.")

        existing_controls = list((await session.execute(select(Control))).scalars().all())
        existing_by_name: dict[str, Control] = {}
        for control in existing_controls:
            normalized_name = _normalize_string(control.name)
            if normalized_name in existing_by_name:
                report.add_error(
                    "duplicate-existing-control-name",
                    (
                        f"Database already contains duplicate control name {control.name!r}. "
                        "Resolve before running import."
                    ),
                    details={"name": control.name},
                )
                continue
            existing_by_name[normalized_name] = control

        matched_risks_by_row: dict[int, list[Risk]] = {}
        planned_links = 0
        for row in rows:
            matched_risks = _match_risks(row.risks_text, all_risks)
            matched_risks_by_row[row.row_number] = matched_risks
            planned_links += len(matched_risks)
            if row.risks_text and not matched_risks:
                report.add_warning(
                    "unmatched-control-risks",
                    f"Control {row.name!r} did not match any risks; it will import without links.",
                    row=row.row_number,
                )

        report.metadata["links_planned"] = planned_links
        if args.allow_reset:
            report.creates = len(rows)
            report.updates = 0
            report.deletes = len(existing_controls)
        else:
            report.creates = sum(1 for row in rows if _normalize_string(row.name) not in existing_by_name)
            report.updates = len(rows) - report.creates

        if not report.ok:
            write_report(args.report, report)
            for issue in report.errors:
                print(f"❌ {issue.code}: {issue.message}")
            return 1

        if not args.apply:
            write_report(args.report, report)
            print("🔍 DRY-RUN: no changes applied.")
            print(
                f"   creates={report.creates} updates={report.updates} deletes={report.deletes} "
                f"links={planned_links} skips={report.skips} warnings={len(report.warnings)}"
            )
            return 0

        if args.allow_reset:
            await session.execute(delete(ControlExecution))
            await session.execute(delete(ControlRiskLink))
            await session.execute(delete(Control))
            await session.flush()
            existing_by_name = {}

        for row in rows:
            normalized_name = _normalize_string(row.name)
            control = existing_by_name.get(normalized_name)

            if control is None:
                control = Control(
                    name=row.name,
                    description=row.description,
                    data_source=row.data_source,
                    methodology_reference=row.methodology_reference,
                    control_form=row.control_form,
                    process_owner_position=row.process_owner_position,
                    control_owner_id=owner.id if owner is not None else None,
                    executor_position=row.executor_position,
                    frequency=row.frequency,
                    risk_level=3,
                    output_description=row.output_description,
                    report_recipient=row.report_recipient,
                    documentation_location=row.documentation_location,
                    department_id=ops_department.id if ops_department is not None else None,
                    status="active",
                    created_by_id=owner.id if owner is not None else None,
                    updated_by_id=owner.id if owner is not None else None,
                )
                session.add(control)
                await session.flush()
                existing_by_name[normalized_name] = control
            else:
                control.name = row.name
                control.description = row.description
                control.data_source = row.data_source
                control.methodology_reference = row.methodology_reference
                control.control_form = row.control_form
                control.process_owner_position = row.process_owner_position
                control.control_owner_id = owner.id if owner is not None else None
                control.executor_position = row.executor_position
                control.frequency = row.frequency
                control.output_description = row.output_description
                control.report_recipient = row.report_recipient
                control.documentation_location = row.documentation_location
                control.department_id = ops_department.id if ops_department is not None else None
                control.status = "active"
                control.updated_by_id = owner.id if owner is not None else None
                await session.execute(delete(ControlRiskLink).where(ControlRiskLink.control_id == control.id))
                await session.flush()

            for risk in matched_risks_by_row[row.row_number]:
                session.add(
                    ControlRiskLink(
                        control_id=control.id,
                        risk_id=risk.id,
                        effectiveness="medium",
                        notes=f"Imported from workbook row {row.row_number}",
                    )
                )

        await session.commit()

    write_report(args.report, report)
    print(
        f"✅ Imported controls: creates={report.creates} updates={report.updates} "
        f"deletes={report.deletes} links={planned_links}"
    )
    return 0


def main() -> int:
    return asyncio.run(_run(_parse_args(sys.argv[1:])))


if __name__ == "__main__":
    raise SystemExit(main())
