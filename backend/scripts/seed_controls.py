"""
Seed script for importing controls from Excel file.
Run with: python -m scripts.seed_controls
"""
import asyncio
import openpyxl
from pathlib import Path
from sqlalchemy import select
from app.db.session import async_session_maker
from app.models import Control, ControlForm, ControlFrequency, ControlStatus, Department, User, Risk

# Column mapping (Czech → English model fields)
COLUMN_MAP = {
    'Pořadové číslo': 'order_number',
    'Název kontroly': 'name',
    'Popis kontroly': 'description',
    'Zdroj dat': 'data_source',
    'Směrnice': 'methodology_reference',
    'Forma': 'control_form',
    'Vlastník procesu': 'process_owner_position',
    'Vlastník kontroly': 'control_owner_dept',  # Will use to assign department
    'Pozice vykonávající kontrolu': 'executor_position',
    'Frekvence': 'frequency',
    'Výstup kontroly': 'output_description',
    'Reportovací povinnost': 'report_recipient',
    'Dokumentace': 'documentation_location',
    'Která rizika snižuje touto kontrolou': 'mitigates_risk',  # For potential linking
    'Nápravné opatření': 'corrective_action',  # Extra info
}

FREQUENCY_MAP = {
    'denně': ControlFrequency.daily.value,
    'týdně': ControlFrequency.weekly.value,
    'měsíčně': ControlFrequency.monthly.value,
    'čtvrtletně': ControlFrequency.quarterly.value,
    'ročně': ControlFrequency.annually.value,
    'ad hoc': ControlFrequency.ad_hoc.value,
}

FORM_MAP = {
    'manuální': ControlForm.manual.value,
    'automatická': ControlForm.automatic.value,
    'automatic': ControlForm.automatic.value,
}


async def get_or_create_department(db, name: str) -> int:
    """Get department by name or create if not exists."""
    result = await db.execute(select(Department).where(Department.name == name))
    dept = result.scalar_one_or_none()
    if dept:
        return dept.id
    
    # Create new department
    new_dept = Department(name=name, code=name[:3].upper())
    db.add(new_dept)
    await db.flush()
    return new_dept.id


async def seed_controls(excel_path: str):
    """Import controls from Excel file."""
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active
    
    # Get header row
    headers = [cell.value for cell in ws[1]]
    
    async with async_session_maker() as db:
        imported_count = 0
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            row_dict = dict(zip(headers, row))
            
            if not row_dict.get('Název kontroly'):
                continue  # Skip empty rows
                
            # Parse frequency
            freq_raw = (row_dict.get('Frekvence') or 'měsíčně').lower().strip()
            frequency = FREQUENCY_MAP.get(freq_raw, ControlFrequency.monthly.value)
            
            # Parse form
            form_raw = (row_dict.get('Forma') or 'manuální').lower().strip()
            control_form = FORM_MAP.get(form_raw, ControlForm.manual.value)
            
            # Get or create department from 'Vlastník kontroly' - use first if comma-separated
            dept_raw = row_dict.get('Vlastník kontroly') or 'Provoz'
            dept_name = dept_raw.split(',')[0].strip()
            department_id = await get_or_create_department(db, dept_name)
            
            control = Control(
                name=row_dict.get('Název kontroly', 'Unknown Control'),
                description=row_dict.get('Popis kontroly') or '',
                data_source=row_dict.get('Zdroj dat'),
                methodology_reference=row_dict.get('Směrnice'),
                control_form=control_form,
                process_owner_position=row_dict.get('Vlastník procesu'),
                executor_position=row_dict.get('Pozice vykonávající kontrolu'),
                frequency=frequency,
                risk_level=3,  # Default medium risk
                output_description=row_dict.get('Výstup kontroly'),
                report_recipient=row_dict.get('Reportovací povinnost'),
                documentation_location=row_dict.get('Dokumentace'),
                department_id=department_id,
                status=ControlStatus.active.value,
            )
            
            db.add(control)
            imported_count += 1
            print(f"Imported: {control.name[:50]}...")
        
        await db.commit()
        print(f"\n✅ Successfully imported {imported_count} controls")


async def seed_mock_risks():
    """Create a few mock risks for testing."""
    mock_risks = [
        {
            'risk_id_code': 'MKT-R01',
            'process': 'Marketing',
            'subprocess': 'Digital Campaigns',
            'risk_type': 'operational',
            'category': 'Reputational',
            'description': 'Risk of reputational damage from negative social media campaigns or customer complaints going viral.',
            'gross_probability': 3, 'gross_impact': 4, 'gross_score': 12,
            'net_probability': 2, 'net_impact': 3, 'net_score': 6,
            'status': 'active', 'is_priority': True,
        },
        {
            'risk_id_code': 'OPS-R01',
            'process': 'Operations',
            'subprocess': 'Policy Processing',
            'risk_type': 'operational',
            'category': 'Process Failure',
            'description': 'Risk of errors in policy processing leading to incorrect coverage or pricing.',
            'gross_probability': 4, 'gross_impact': 3, 'gross_score': 12,
            'net_probability': 2, 'net_impact': 2, 'net_score': 4,
            'status': 'active', 'is_priority': False,
        },
        {
            'risk_id_code': 'FIN-R01',
            'process': 'Finance',
            'subprocess': 'Claims Settlement',
            'risk_type': 'strategic',
            'category': 'Financial',
            'description': 'Risk of fraudulent claims resulting in financial losses.',
            'gross_probability': 3, 'gross_impact': 5, 'gross_score': 15,
            'net_probability': 2, 'net_impact': 4, 'net_score': 8,
            'status': 'monitoring', 'is_priority': True,
        },
        {
            'risk_id_code': 'IT-R01',
            'process': 'IT & Security',
            'subprocess': 'Cybersecurity',
            'risk_type': 'operational',
            'category': 'Cyber',
            'description': 'Risk of data breach or ransomware attack affecting customer data and operations.',
            'gross_probability': 4, 'gross_impact': 5, 'gross_score': 20,
            'net_probability': 2, 'net_impact': 4, 'net_score': 8,
            'status': 'active', 'is_priority': True,
        },
        {
            'risk_id_code': 'HR-R01',
            'process': 'Human Resources',
            'subprocess': 'Talent Retention',
            'risk_type': 'strategic',
            'category': 'People',
            'description': 'Risk of losing key personnel to competitors affecting business continuity.',
            'gross_probability': 3, 'gross_impact': 3, 'gross_score': 9,
            'net_probability': 2, 'net_impact': 2, 'net_score': 4,
            'status': 'active', 'is_priority': False,
        },
    ]
    
    async with async_session_maker() as db:
        for risk_data in mock_risks:
            # Check if already exists
            result = await db.execute(
                select(Risk).where(Risk.risk_id_code == risk_data['risk_id_code'])
            )
            if result.scalar_one_or_none():
                print(f"Risk {risk_data['risk_id_code']} already exists, skipping...")
                continue
            
            risk = Risk(**risk_data)
            db.add(risk)
            print(f"Created risk: {risk_data['risk_id_code']}")
        
        await db.commit()
        print("\n✅ Mock risks seeded")


async def main():
    excel_path = Path(__file__).parent.parent.parent / 'placeholder-controls-source.xlsx'
    
    print("🌱 Seeding database...")
    print(f"📁 Excel file: {excel_path}")
    
    await seed_controls(str(excel_path))
    await seed_mock_risks()
    
    print("\n🎉 Seeding complete!")


if __name__ == "__main__":
    asyncio.run(main())
