from __future__ import annotations

from io import BytesIO
from typing import TYPE_CHECKING

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

if TYPE_CHECKING:
    from app.models.control import Control


def generate_controls_excel(controls: list["Control"]) -> bytes:
    """Generate an Excel report of controls."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Controls"

    # Header styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1e293b", end_color="1e293b", fill_type="solid")
    header_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Headers
    headers = [
        "ID",
        "Name",
        "Description",
        "Department",
        "Status",
        "Form",
        "Frequency",
        "Risk Level",
        "Owner Position",
        "Executor Position",
        "Data Source",
        "Methodology Reference",
        "Created At",
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = header_border
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    for row, control in enumerate(controls, 2):
        ws.cell(row=row, column=1, value=control.id)
        ws.cell(row=row, column=2, value=control.name)
        ws.cell(row=row, column=3, value=control.description[:500] if control.description else "")
        ws.cell(row=row, column=4, value=control.department.name if control.department else "N/A")
        ws.cell(row=row, column=5, value=control.status)
        ws.cell(row=row, column=6, value=control.control_form)
        ws.cell(row=row, column=7, value=control.frequency)
        ws.cell(row=row, column=8, value=control.risk_level)
        ws.cell(row=row, column=9, value=control.process_owner_position or "")
        ws.cell(row=row, column=10, value=control.executor_position or "")
        ws.cell(row=row, column=11, value=control.data_source or "")
        ws.cell(row=row, column=12, value=control.methodology_reference or "")
        ws.cell(row=row, column=13, value=control.created_at.strftime("%Y-%m-%d") if control.created_at else "")

    # Adjust column widths
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 50
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 12
    ws.column_dimensions["H"].width = 12
    ws.column_dimensions["I"].width = 20
    ws.column_dimensions["J"].width = 20
    ws.column_dimensions["K"].width = 25
    ws.column_dimensions["L"].width = 25
    ws.column_dimensions["M"].width = 15

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()

