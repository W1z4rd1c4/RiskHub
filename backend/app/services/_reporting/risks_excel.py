from __future__ import annotations

from io import BytesIO
from typing import TYPE_CHECKING

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

if TYPE_CHECKING:
    from app.models.risk import Risk


def generate_risks_excel(risks: list["Risk"]) -> bytes:
    """Generate an Excel report of risks."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Risks"

    # Header styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1e293b", end_color="1e293b", fill_type="solid")
    header_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Headers
    headers = [
        "ID",
        "Risk ID Code",
        "Process",
        "Description",
        "Category",
        "Department",
        "Status",
        "Gross Probability",
        "Gross Impact",
        "Gross Score",
        "Net Probability",
        "Net Impact",
        "Net Score",
        "Owner",
        "Treatment Strategy",
        "Created At",
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = header_border
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    for row, risk in enumerate(risks, 2):
        ws.cell(row=row, column=1, value=risk.id)
        ws.cell(row=row, column=2, value=risk.risk_id_code)
        ws.cell(row=row, column=3, value=risk.process)
        ws.cell(row=row, column=4, value=risk.description[:500] if risk.description else "")
        ws.cell(row=row, column=5, value=risk.category or "")
        ws.cell(row=row, column=6, value=risk.department.name if risk.department else "N/A")
        ws.cell(row=row, column=7, value=risk.status or "")
        ws.cell(row=row, column=8, value=risk.gross_probability)
        ws.cell(row=row, column=9, value=risk.gross_impact)
        ws.cell(row=row, column=10, value=risk.gross_probability * risk.gross_impact)
        ws.cell(row=row, column=11, value=risk.net_probability)
        ws.cell(row=row, column=12, value=risk.net_impact)
        ws.cell(row=row, column=13, value=risk.net_probability * risk.net_impact)
        ws.cell(row=row, column=14, value=risk.owner.name if hasattr(risk, "owner") and risk.owner else "")
        ws.cell(row=row, column=15, value="")  # Treatment strategy not in model
        ws.cell(row=row, column=16, value=risk.created_at.strftime("%Y-%m-%d") if risk.created_at else "")

    # Adjust column widths
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 50
    ws.column_dimensions["D"].width = 15
    ws.column_dimensions["E"].width = 20
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 15
    ws.column_dimensions["H"].width = 12
    ws.column_dimensions["I"].width = 12
    ws.column_dimensions["J"].width = 14
    ws.column_dimensions["K"].width = 12
    ws.column_dimensions["L"].width = 12
    ws.column_dimensions["M"].width = 20
    ws.column_dimensions["N"].width = 18
    ws.column_dimensions["O"].width = 15

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()

