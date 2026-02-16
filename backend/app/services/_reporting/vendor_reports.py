from __future__ import annotations

from io import BytesIO
from typing import TYPE_CHECKING

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

if TYPE_CHECKING:
    from app.schemas.vendor_reports import VendorAnnualReportData, VendorDoraRegisterRow


def generate_vendor_annual_report_excel(report: "VendorAnnualReportData") -> bytes:
    wb = Workbook()

    ws = wb.active
    ws.title = "Vendors"

    headers = [
        "Vendor ID",
        "Name",
        "Legal Name",
        "Vendor Type",
        "Department",
        "Owner",
        "Process",
        "Subprocess",
        "Supports Core Function",
        "DORA Relevant",
        "Significant Vendor",
        "Risk Score (1-5)",
        "Last Decision",
        "Next Reassessment Due",
        "Cadence (months)",
        "Major Breaches (count)",
        "Major Incidents (count)",
        "Major Items (preview)",
    ]

    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="1e293b", end_color="1e293b", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    for v in report.vendors:
        ws.append(
            [
                v.vendor_id,
                v.name,
                v.legal_name or "",
                v.vendor_type,
                v.department_name or "",
                v.outsourcing_owner_name or "",
                v.process,
                v.subprocess or "",
                bool(v.supports_important_core_insurance_function),
                bool(v.dora_relevant),
                bool(v.is_significant_vendor),
                v.risk_score_1_5,
                v.last_decided_at.isoformat() if v.last_decided_at else "",
                v.next_reassessment_due_at.isoformat() if v.next_reassessment_due_at else "",
                v.reassessment_cadence_months,
                v.major_breaches_count,
                v.major_incidents_count,
                "; ".join(v.major_items_preview or []),
            ]
        )

    ws2 = wb.create_sheet("Process Evaluation")
    ws2.append(["Year", report.process_evaluation.year])
    ws2.append(["Generated At", report.generated_at.isoformat()])
    ws2.append(["Total Active Vendors", report.process_evaluation.total_active_vendors])
    ws2.append(["Overdue Reassessments", report.process_evaluation.overdue_reassessments_count])
    ws2.append(["Missing Exit Plans", report.process_evaluation.missing_exit_plans_count])
    ws2.append(["Missing Contingency Plans", report.process_evaluation.missing_contingency_plans_count])
    ws2.append(["Major Breaches (count)", report.process_evaluation.major_breaches_count])
    ws2.append(["Major Incidents (count)", report.process_evaluation.major_incidents_count])

    for row in ws2.iter_rows(min_row=1, max_col=2):
        for cell in row:
            cell.border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def generate_vendor_dora_register_excel(rows: list["VendorDoraRegisterRow"]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "DORA Register"

    headers = [
        "vendor_id",
        "name",
        "legal_name",
        "registration_id",
        "vendor_type",
        "dora_relevant",
        "is_significant_vendor",
        "supports_important_core_insurance_function",
        "risk_score_1_5",
        "outsourcing_owner_user_id",
        "outsourcing_owner_name",
        "department_id",
        "department_name",
        "process",
        "subprocess",
        "last_decided_at",
        "next_reassessment_due_at",
        "reassessment_cadence_months",
        "replaceability",
        "has_alternative_providers",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="1e293b", end_color="1e293b", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    for r in rows:
        ws.append(
            [
                r.vendor_id,
                r.name,
                r.legal_name or "",
                r.registration_id or "",
                r.vendor_type,
                bool(r.dora_relevant),
                bool(r.is_significant_vendor),
                bool(r.supports_important_core_insurance_function),
                r.risk_score_1_5,
                r.outsourcing_owner_user_id or "",
                r.outsourcing_owner_name or "",
                r.department_id or "",
                r.department_name or "",
                r.process,
                r.subprocess or "",
                r.last_decided_at.isoformat() if r.last_decided_at else "",
                r.next_reassessment_due_at.isoformat() if r.next_reassessment_due_at else "",
                r.reassessment_cadence_months,
                r.replaceability or "",
                bool(r.has_alternative_providers),
            ]
        )

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()

