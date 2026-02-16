from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


def generate_audit_trail_excel(executions: list) -> bytes:
    """Generate an Excel report of control executions for audit trail."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Audit Trail"

    # Header styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1e293b", end_color="1e293b", fill_type="solid")
    header_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    wrap_alignment = Alignment(wrap_text=True, vertical="top")

    # Headers
    headers = [
        "ID",
        "Executed At",
        "Control ID",
        "Control Name",
        "Department",
        "Executor",
        "Result",
        "Findings",
        "Evidence Reference",
        "Notes",
        "Next Scheduled",
        "Linked Risks",
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = header_border
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    for row, exe in enumerate(executions, 2):
        ws.cell(row=row, column=1, value=exe.id)
        ws.cell(row=row, column=2, value=exe.executed_at.strftime("%Y-%m-%d %H:%M") if exe.executed_at else "")
        ws.cell(row=row, column=3, value=exe.control_id)
        ws.cell(row=row, column=4, value=exe.control.name if exe.control else "")
        ws.cell(row=row, column=5, value=exe.control.department.name if exe.control and exe.control.department else "")
        ws.cell(row=row, column=6, value=exe.executed_by.name if exe.executed_by else "")
        ws.cell(row=row, column=7, value=exe.result or "")

        # Findings with wrap
        findings_cell = ws.cell(row=row, column=8, value=exe.findings or "")
        findings_cell.alignment = wrap_alignment

        # Evidence with wrap
        evidence_cell = ws.cell(row=row, column=9, value=exe.evidence_reference or "")
        evidence_cell.alignment = wrap_alignment

        # Notes with wrap
        notes_cell = ws.cell(row=row, column=10, value=exe.notes or "")
        notes_cell.alignment = wrap_alignment

        ws.cell(row=row, column=11, value=exe.next_scheduled.strftime("%Y-%m-%d") if exe.next_scheduled else "")

        # Linked risks (from control.risk_links -> risk)
        linked_risks = ""
        if exe.control and hasattr(exe.control, "risk_links"):
            risk_names = []
            for link in exe.control.risk_links:
                if hasattr(link, "risk") and link.risk:
                    display_name = (link.risk.name or link.risk.process or "").strip()
                    if display_name:
                        risk_names.append(f"R-{link.risk.id}: {display_name[:30]}")
                    else:
                        risk_names.append(f"R-{link.risk.id}")
            linked_risks = "; ".join(risk_names)
        ws.cell(row=row, column=12, value=linked_risks)

    # Adjust column widths
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 30
    ws.column_dimensions["E"].width = 20
    ws.column_dimensions["F"].width = 18
    ws.column_dimensions["G"].width = 12
    ws.column_dimensions["H"].width = 40
    ws.column_dimensions["I"].width = 30
    ws.column_dimensions["J"].width = 30
    ws.column_dimensions["K"].width = 15
    ws.column_dimensions["L"].width = 40

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()

