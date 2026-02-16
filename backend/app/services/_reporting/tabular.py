from __future__ import annotations

import csv
from datetime import date, datetime
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


def _format_export_value(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, date):
        return value.isoformat()
    return str(value)


def _sanitize_csv_cell(value: str) -> str:
    """
    Prevent CSV formula injection by prefixing risky leading characters.
    """
    if not value:
        return value
    if value[0] in ("=", "+", "-", "@", "\t", "\r"):
        return f"'{value}"
    return value


def generate_tabular_csv(headers: list[str], rows: list[list[object]]) -> bytes:
    buffer = BytesIO()
    text_buffer = []

    class _Writer:
        def write(self, chunk: str) -> int:
            text_buffer.append(chunk)
            return len(chunk)

    writer = csv.writer(_Writer())
    writer.writerow(headers)
    for row in rows:
        writer.writerow([_sanitize_csv_cell(_format_export_value(cell)) for cell in row])

    buffer.write("".join(text_buffer).encode("utf-8"))
    buffer.seek(0)
    return buffer.getvalue()


def generate_tabular_excel(sheet_name: str, headers: list[str], rows: list[list[object]]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31] if sheet_name else "Export"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1e293b", end_color="1e293b", fill_type="solid")
    header_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = header_border
        cell.alignment = Alignment(horizontal="center")

    max_len = [len(h) for h in headers]
    for row_idx, row in enumerate(rows, start=2):
        for col_idx, cell_value in enumerate(row, start=1):
            value = _format_export_value(cell_value)
            ws.cell(row=row_idx, column=col_idx, value=value)
            max_len[col_idx - 1] = max(max_len[col_idx - 1], len(value))

    for i, width in enumerate(max_len, start=1):
        # Cap width to keep workbook readable
        ws.column_dimensions[chr(64 + i)].width = min(max(width + 2, 10), 60)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()

