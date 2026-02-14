"""
Report generation service for Excel and CSV exports.
"""
import csv
from datetime import date, datetime
from io import BytesIO
from typing import TYPE_CHECKING

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

if TYPE_CHECKING:
    from app.models.control import Control
    from app.models.risk import Risk
    from app.schemas.vendor_reports import VendorAnnualReportData, VendorDoraRegisterRow


def count_high_risks(risks, high_threshold: int) -> int:
    """Count risks with net_probability * net_impact >= high_threshold.

    Used by threshold propagation tests and report summaries.
    """
    total = 0
    for r in risks:
        prob = getattr(r, "net_probability", 0) or 0
        impact = getattr(r, "net_impact", 0) or 0
        if (prob * impact) >= high_threshold:
            total += 1
    return total


def generate_controls_excel(controls: list["Control"]) -> bytes:
    """Generate an Excel report of controls."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Controls"

    # Header styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1e293b", end_color="1e293b", fill_type="solid")
    header_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Headers
    headers = [
        "ID", "Name", "Description", "Department", "Status", "Form",
        "Frequency", "Risk Level", "Owner Position", "Executor Position",
        "Data Source", "Methodology Reference", "Created At"
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = header_border
        cell.alignment = Alignment(horizontal='center')

    # Data rows
    for row, control in enumerate(controls, 2):
        ws.cell(row=row, column=1, value=control.id)
        ws.cell(row=row, column=2, value=control.name)
        ws.cell(row=row, column=3, value=control.description[:500] if control.description else "")
        ws.cell(row=row, column=4, value=control.department.name if control.department else "N/A")
        ws.cell(row=row, column=5, value=control.status)
        ws.cell(row=row, column=6, value=control.control_form)
        ws.cell(row=row, column=7, value=control.frequency)
        ws.cell(row=row, column=8, value=control.risk_level)
        ws.cell(row=row, column=9, value=control.process_owner_position or "")
        ws.cell(row=row, column=10, value=control.executor_position or "")
        ws.cell(row=row, column=11, value=control.data_source or "")
        ws.cell(row=row, column=12, value=control.methodology_reference or "")
        ws.cell(row=row, column=13, value=control.created_at.strftime('%Y-%m-%d') if control.created_at else "")

    # Adjust column widths
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 50
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 12
    ws.column_dimensions['I'].width = 20
    ws.column_dimensions['J'].width = 20
    ws.column_dimensions['K'].width = 25
    ws.column_dimensions['L'].width = 25
    ws.column_dimensions['M'].width = 15

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


# =============================================================================
# Vendor Reports (Phase 18-09)
# =============================================================================


def generate_vendor_annual_report_excel(report: "VendorAnnualReportData") -> bytes:
    wb = Workbook()

    ws = wb.active
    ws.title = "Vendors"

    headers = [
        "Vendor ID",
        "Name",
        "Legal Name",
        "Vendor Type",
        "Department",
        "Owner",
        "Process",
        "Subprocess",
        "Supports Core Function",
        "DORA Relevant",
        "Significant Vendor",
        "Risk Score (1-5)",
        "Last Decision",
        "Next Reassessment Due",
        "Cadence (months)",
        "Major Breaches (count)",
        "Major Incidents (count)",
        "Major Items (preview)",
    ]

    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="1e293b", end_color="1e293b", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    for v in report.vendors:
        ws.append(
            [
                v.vendor_id,
                v.name,
                v.legal_name or "",
                v.vendor_type,
                v.department_name or "",
                v.outsourcing_owner_name or "",
                v.process,
                v.subprocess or "",
                bool(v.supports_important_core_insurance_function),
                bool(v.dora_relevant),
                bool(v.is_significant_vendor),
                v.risk_score_1_5,
                v.last_decided_at.isoformat() if v.last_decided_at else "",
                v.next_reassessment_due_at.isoformat() if v.next_reassessment_due_at else "",
                v.reassessment_cadence_months,
                v.major_breaches_count,
                v.major_incidents_count,
                "; ".join(v.major_items_preview or []),
            ]
        )

    ws2 = wb.create_sheet("Process Evaluation")
    ws2.append(["Year", report.process_evaluation.year])
    ws2.append(["Generated At", report.generated_at.isoformat()])
    ws2.append(["Total Active Vendors", report.process_evaluation.total_active_vendors])
    ws2.append(["Overdue Reassessments", report.process_evaluation.overdue_reassessments_count])
    ws2.append(["Missing Exit Plans", report.process_evaluation.missing_exit_plans_count])
    ws2.append(["Missing Contingency Plans", report.process_evaluation.missing_contingency_plans_count])
    ws2.append(["Major Breaches (count)", report.process_evaluation.major_breaches_count])
    ws2.append(["Major Incidents (count)", report.process_evaluation.major_incidents_count])

    for row in ws2.iter_rows(min_row=1, max_col=2):
        for cell in row:
            cell.border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def generate_vendor_dora_register_excel(rows: list["VendorDoraRegisterRow"]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "DORA Register"

    headers = [
        "vendor_id",
        "name",
        "legal_name",
        "registration_id",
        "vendor_type",
        "dora_relevant",
        "is_significant_vendor",
        "supports_important_core_insurance_function",
        "risk_score_1_5",
        "outsourcing_owner_user_id",
        "outsourcing_owner_name",
        "department_id",
        "department_name",
        "process",
        "subprocess",
        "last_decided_at",
        "next_reassessment_due_at",
        "reassessment_cadence_months",
        "replaceability",
        "has_alternative_providers",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="1e293b", end_color="1e293b", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    for r in rows:
        ws.append(
            [
                r.vendor_id,
                r.name,
                r.legal_name or "",
                r.registration_id or "",
                r.vendor_type,
                bool(r.dora_relevant),
                bool(r.is_significant_vendor),
                bool(r.supports_important_core_insurance_function),
                r.risk_score_1_5,
                r.outsourcing_owner_user_id or "",
                r.outsourcing_owner_name or "",
                r.department_id or "",
                r.department_name or "",
                r.process,
                r.subprocess or "",
                r.last_decided_at.isoformat() if r.last_decided_at else "",
                r.next_reassessment_due_at.isoformat() if r.next_reassessment_due_at else "",
                r.reassessment_cadence_months,
                r.replaceability or "",
                bool(r.has_alternative_providers),
            ]
        )

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def generate_risks_excel(risks: list["Risk"]) -> bytes:
    """Generate an Excel report of risks."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Risks"

    # Header styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1e293b", end_color="1e293b", fill_type="solid")
    header_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Headers
    headers = [
        "ID", "Risk ID Code", "Process", "Description", "Category", "Department", "Status",
        "Gross Probability", "Gross Impact", "Gross Score",
        "Net Probability", "Net Impact", "Net Score",
        "Owner", "Treatment Strategy", "Created At"
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = header_border
        cell.alignment = Alignment(horizontal='center')

    # Data rows
    for row, risk in enumerate(risks, 2):
        ws.cell(row=row, column=1, value=risk.id)
        ws.cell(row=row, column=2, value=risk.risk_id_code)
        ws.cell(row=row, column=3, value=risk.process)
        ws.cell(row=row, column=4, value=risk.description[:500] if risk.description else "")
        ws.cell(row=row, column=5, value=risk.category or "")
        ws.cell(row=row, column=6, value=risk.department.name if risk.department else "N/A")
        ws.cell(row=row, column=7, value=risk.status or "")
        ws.cell(row=row, column=8, value=risk.gross_probability)
        ws.cell(row=row, column=9, value=risk.gross_impact)
        ws.cell(row=row, column=10, value=risk.gross_probability * risk.gross_impact)
        ws.cell(row=row, column=11, value=risk.net_probability)
        ws.cell(row=row, column=12, value=risk.net_impact)
        ws.cell(row=row, column=13, value=risk.net_probability * risk.net_impact)
        ws.cell(row=row, column=14, value=risk.owner.name if hasattr(risk, 'owner') and risk.owner else "")
        ws.cell(row=row, column=15, value="")  # Treatment strategy not in model
        ws.cell(row=row, column=16, value=risk.created_at.strftime('%Y-%m-%d') if risk.created_at else "")

    # Adjust column widths
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 50
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 12
    ws.column_dimensions['I'].width = 12
    ws.column_dimensions['J'].width = 14
    ws.column_dimensions['K'].width = 12
    ws.column_dimensions['L'].width = 12
    ws.column_dimensions['M'].width = 20
    ws.column_dimensions['N'].width = 18
    ws.column_dimensions['O'].width = 15

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def generate_audit_trail_excel(executions: list) -> bytes:
    """Generate an Excel report of control executions for audit trail."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Audit Trail"

    # Header styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1e293b", end_color="1e293b", fill_type="solid")
    header_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    wrap_alignment = Alignment(wrap_text=True, vertical='top')

    # Headers
    headers = [
        "ID", "Executed At", "Control ID", "Control Name", "Department",
        "Executor", "Result", "Findings", "Evidence Reference", "Notes",
        "Next Scheduled", "Linked Risks"
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = header_border
        cell.alignment = Alignment(horizontal='center')

    # Data rows
    for row, exe in enumerate(executions, 2):
        ws.cell(row=row, column=1, value=exe.id)
        ws.cell(row=row, column=2, value=exe.executed_at.strftime('%Y-%m-%d %H:%M') if exe.executed_at else "")
        ws.cell(row=row, column=3, value=exe.control_id)
        ws.cell(row=row, column=4, value=exe.control.name if exe.control else "")
        ws.cell(row=row, column=5, value=exe.control.department.name if exe.control and exe.control.department else "")
        ws.cell(row=row, column=6, value=exe.executed_by.name if exe.executed_by else "")
        ws.cell(row=row, column=7, value=exe.result or "")

        # Findings with wrap
        findings_cell = ws.cell(row=row, column=8, value=exe.findings or "")
        findings_cell.alignment = wrap_alignment

        # Evidence with wrap
        evidence_cell = ws.cell(row=row, column=9, value=exe.evidence_reference or "")
        evidence_cell.alignment = wrap_alignment

        # Notes with wrap
        notes_cell = ws.cell(row=row, column=10, value=exe.notes or "")
        notes_cell.alignment = wrap_alignment

        ws.cell(row=row, column=11, value=exe.next_scheduled.strftime('%Y-%m-%d') if exe.next_scheduled else "")

        # Linked risks (from control.risk_links -> risk)
        linked_risks = ""
        if exe.control and hasattr(exe.control, 'risk_links'):
            risk_names = []
            for link in exe.control.risk_links:
                if hasattr(link, 'risk') and link.risk:
                    display_name = (link.risk.name or link.risk.process or "").strip()
                    if display_name:
                        risk_names.append(f"R-{link.risk.id}: {display_name[:30]}")
                    else:
                        risk_names.append(f"R-{link.risk.id}")
            linked_risks = "; ".join(risk_names)
        ws.cell(row=row, column=12, value=linked_risks)

    # Adjust column widths
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 30
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 18
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 40
    ws.column_dimensions['I'].width = 30
    ws.column_dimensions['J'].width = 30
    ws.column_dimensions['K'].width = 15
    ws.column_dimensions['L'].width = 40

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def _format_export_value(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, date):
        return value.isoformat()
    return str(value)


def _sanitize_csv_cell(value: str) -> str:
    """
    Prevent CSV formula injection by prefixing risky leading characters.
    """
    if not value:
        return value
    if value[0] in ("=", "+", "-", "@", "\t", "\r"):
        return f"'{value}"
    return value


def generate_tabular_csv(headers: list[str], rows: list[list[object]]) -> bytes:
    buffer = BytesIO()
    text_buffer = []

    class _Writer:
        def write(self, chunk: str) -> int:
            text_buffer.append(chunk)
            return len(chunk)

    writer = csv.writer(_Writer())
    writer.writerow(headers)
    for row in rows:
        writer.writerow([_sanitize_csv_cell(_format_export_value(cell)) for cell in row])

    buffer.write("".join(text_buffer).encode("utf-8"))
    buffer.seek(0)
    return buffer.getvalue()


def generate_tabular_excel(sheet_name: str, headers: list[str], rows: list[list[object]]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31] if sheet_name else "Export"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1e293b", end_color="1e293b", fill_type="solid")
    header_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = header_border
        cell.alignment = Alignment(horizontal="center")

    max_len = [len(h) for h in headers]
    for row_idx, row in enumerate(rows, start=2):
        for col_idx, cell_value in enumerate(row, start=1):
            value = _format_export_value(cell_value)
            ws.cell(row=row_idx, column=col_idx, value=value)
            max_len[col_idx - 1] = max(max_len[col_idx - 1], len(value))

    for i, width in enumerate(max_len, start=1):
        # Cap width to keep workbook readable
        ws.column_dimensions[chr(64 + i)].width = min(max(width + 2, 10), 60)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()
