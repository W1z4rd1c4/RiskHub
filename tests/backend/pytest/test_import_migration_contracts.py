from __future__ import annotations

import argparse
import json
from contextlib import asynccontextmanager
from pathlib import Path

import openpyxl
import pytest
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models import Control, ControlRiskLink, Department, KeyRiskIndicator, Risk
from scripts import migrate_controls, migrate_kris, migrate_risks


def _write_risk_workbook(path: Path, rows: list[dict[str, object]]) -> None:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Rizika"

    for offset, row in enumerate(rows, start=9):
        sheet.cell(row=offset, column=2).value = row["process"]
        sheet.cell(row=offset, column=3).value = row.get("subprocess")
        sheet.cell(row=offset, column=4).value = row.get("risk_type", "Operační riziko")
        sheet.cell(row=offset, column=6).value = row["name"]
        sheet.cell(row=offset, column=7).value = row.get("description", "")
        sheet.cell(row=offset, column=8).value = row.get("gross_impact", 3)
        sheet.cell(row=offset, column=9).value = row.get("gross_probability", 3)
        sheet.cell(row=offset, column=12).value = row.get("net_impact", 2)
        sheet.cell(row=offset, column=13).value = row.get("net_probability", 2)

    workbook.save(path)


def _write_controls_workbook(path: Path) -> None:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.cell(row=2, column=2).value = "Quarterly Finance Review"
    sheet.cell(row=2, column=3).value = "Updated control description"
    sheet.cell(row=2, column=4).value = "Ledger export"
    sheet.cell(row=2, column=5).value = "FIN-CTRL-001"
    sheet.cell(row=2, column=6).value = "Manuální"
    sheet.cell(row=2, column=7).value = "Finance Director"
    sheet.cell(row=2, column=9).value = "Controller"
    sheet.cell(row=2, column=10).value = "měsíčně"
    sheet.cell(row=2, column=11).value = "Review notes"
    sheet.cell(row=2, column=12).value = "Audit committee"
    sheet.cell(row=2, column=13).value = "SharePoint"
    sheet.cell(row=2, column=14).value = "Late reconciliations on core finance process"
    sheet.cell(row=2, column=15).value = "Escalate repeated breaches"
    workbook.save(path)


def _write_kri_workbook(path: Path) -> None:
    workbook = openpyxl.Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)
    sheet = workbook.create_sheet("Provozní riziko")
    sheet.cell(row=2, column=2).value = "Unknown Process"
    sheet.cell(row=2, column=6).value = "zxqv sentinel unmapped narrative"
    sheet.cell(row=2, column=9).value = "Alien backlog %"
    sheet.cell(row=2, column=10).value = 12
    sheet.cell(row=2, column=11).value = 1
    sheet.cell(row=2, column=12).value = 5
    workbook.save(path)


def _risk_fixture(
    *,
    risk_id_code: str,
    name: str,
    process: str,
    subprocess: str | None,
    description: str,
    department_id: int,
    owner_id: int,
) -> Risk:
    return Risk(
        risk_id_code=risk_id_code,
        name=name,
        process=process,
        subprocess=subprocess,
        risk_type="operational",
        category="Operational",
        description=description,
        department_id=department_id,
        owner_id=owner_id,
        gross_probability=3,
        gross_impact=3,
        gross_score=9,
        net_probability=2,
        net_impact=2,
        net_score=4,
        status="active",
        is_priority=False,
    )


def _patch_script_session(
    monkeypatch: pytest.MonkeyPatch,
    module,
    db_session: AsyncSession,
) -> None:
    @asynccontextmanager
    async def fake_session_context(_settings):
        yield db_session

    monkeypatch.setattr(module, "session_context", fake_session_context)
    monkeypatch.setattr(module, "get_settings", lambda: None)


@pytest.mark.asyncio
async def test_migrate_risks_dry_run_is_non_destructive_and_reports_identity_mode(
    db_session: AsyncSession,
    test_department: Department,
    test_user,
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
):
    existing_risk = _risk_fixture(
        risk_id_code="MARK-R01",
        name="Liquidity Label",
        process="Marketing",
        subprocess="Campaign approvals",
        description="Existing detailed narrative",
        department_id=test_department.id,
        owner_id=test_user.id,
    )
    db_session.add(existing_risk)
    await db_session.commit()

    workbook_path = tmp_path / "risks.xlsx"
    report_path = tmp_path / "risk-report.json"
    _write_risk_workbook(
        workbook_path,
        [
            {
                "process": "Marketing",
                "subprocess": "Campaign approvals",
                "risk_type": "Strategické",
                "name": "Liquidity Label",
                "description": "Workbook narrative should not apply in dry-run",
                "gross_impact": 4,
                "gross_probability": 5,
                "net_impact": 2,
                "net_probability": 3,
            }
        ],
    )

    _patch_script_session(monkeypatch, migrate_risks, db_session)

    rc = await migrate_risks._run(
        argparse.Namespace(
            input=str(workbook_path),
            apply=False,
            allow_reset=False,
            report=str(report_path),
        )
    )

    assert rc == 0
    refreshed = (await db_session.execute(select(Risk).where(Risk.id == existing_risk.id))).scalar_one()
    assert refreshed.risk_id_code == "MARK-R01"
    assert refreshed.description == "Existing detailed narrative"

    report = json.loads(report_path.read_text(encoding="utf-8"))
    assert report["creates"] == 0
    assert report["updates"] == 1
    assert report["deletes"] == 0
    assert report["metadata"]["identity_mode"] == "process_subprocess_name"
    assert report["metadata"]["matched_count"] == 1
    assert report["metadata"]["created_count"] == 0


@pytest.mark.asyncio
async def test_migrate_risks_non_reset_preserves_ids_codes_and_linked_records_when_rows_are_reordered(
    db_session: AsyncSession,
    test_department: Department,
    test_user,
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
):
    risk_a = _risk_fixture(
        risk_id_code="CLAI-R01",
        name="Claims Queue Label",
        process="Claims",
        subprocess="Queue",
        description="Original queue narrative",
        department_id=test_department.id,
        owner_id=test_user.id,
    )
    risk_b = _risk_fixture(
        risk_id_code="CLAI-R02",
        name="Claims Reserve Label",
        process="Claims",
        subprocess="Reserve",
        description="Original reserve narrative",
        department_id=test_department.id,
        owner_id=test_user.id,
    )
    control = Control(
        name="Claims Reserve Review",
        description="Reserve review",
        control_form="manual",
        frequency="monthly",
        risk_level=3,
        department_id=test_department.id,
        status="active",
        control_owner_id=test_user.id,
        created_by_id=test_user.id,
        updated_by_id=test_user.id,
    )
    db_session.add_all([risk_a, risk_b, control])
    await db_session.commit()
    await db_session.refresh(risk_a)
    await db_session.refresh(risk_b)
    await db_session.refresh(control)

    kri = KeyRiskIndicator(
        risk_id=risk_a.id,
        metric_name="Queue breach %",
        description="Queue metric",
        current_value=5.0,
        lower_limit=0.0,
        upper_limit=10.0,
        unit="%",
    )
    link = ControlRiskLink(control_id=control.id, risk_id=risk_b.id, effectiveness="medium")
    db_session.add_all([kri, link])
    await db_session.commit()

    workbook_path = tmp_path / "risks-reordered.xlsx"
    report_path = tmp_path / "risk-reordered-report.json"
    _write_risk_workbook(
        workbook_path,
        [
            {
                "process": "Claims",
                "subprocess": "Reserve",
                "risk_type": "Operační riziko",
                "name": "Claims Reserve Label",
                "description": "Updated reserve narrative from column G",
            },
            {
                "process": "Claims",
                "subprocess": "Queue",
                "risk_type": "Operační riziko",
                "name": "Claims Queue Label",
                "description": "Updated queue narrative from column G",
            },
        ],
    )

    _patch_script_session(monkeypatch, migrate_risks, db_session)

    rc = await migrate_risks._run(
        argparse.Namespace(
            input=str(workbook_path),
            apply=True,
            allow_reset=False,
            report=str(report_path),
        )
    )

    assert rc == 0
    refreshed_a = (await db_session.execute(select(Risk).where(Risk.id == risk_a.id))).scalar_one()
    refreshed_b = (await db_session.execute(select(Risk).where(Risk.id == risk_b.id))).scalar_one()
    assert refreshed_a.risk_id_code == "CLAI-R01"
    assert refreshed_b.risk_id_code == "CLAI-R02"
    assert refreshed_a.description == "Updated queue narrative from column G"
    assert refreshed_b.description == "Updated reserve narrative from column G"
    assert refreshed_a.name == "Claims Queue Label"
    assert refreshed_b.name == "Claims Reserve Label"

    persisted_kri = (await db_session.execute(select(KeyRiskIndicator))).scalar_one()
    persisted_link = (await db_session.execute(select(ControlRiskLink))).scalar_one()
    assert persisted_kri.risk_id == risk_a.id
    assert persisted_link.risk_id == risk_b.id

    report = json.loads(report_path.read_text(encoding="utf-8"))
    assert report["creates"] == 0
    assert report["updates"] == 2
    assert report["metadata"]["matched_count"] == 2
    assert report["metadata"]["created_count"] == 0


@pytest.mark.asyncio
async def test_migrate_risks_non_reset_creates_new_risk_without_renumbering_existing_process_rows(
    db_session: AsyncSession,
    test_department: Department,
    test_user,
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
):
    existing_risk = _risk_fixture(
        risk_id_code="CLAI-R01",
        name="Claims Existing Label",
        process="Claims",
        subprocess="Queue",
        description="Existing queue narrative",
        department_id=test_department.id,
        owner_id=test_user.id,
    )
    db_session.add(existing_risk)
    await db_session.commit()
    await db_session.refresh(existing_risk)

    workbook_path = tmp_path / "risks-inserted.xlsx"
    report_path = tmp_path / "risk-inserted-report.json"
    _write_risk_workbook(
        workbook_path,
        [
            {
                "process": "Claims",
                "subprocess": "Alerts",
                "name": "Claims New Label",
                "description": "New alert narrative",
            },
            {
                "process": "Claims",
                "subprocess": "Queue",
                "name": "Claims Existing Label",
                "description": "Updated queue narrative",
            },
        ],
    )

    _patch_script_session(monkeypatch, migrate_risks, db_session)

    rc = await migrate_risks._run(
        argparse.Namespace(
            input=str(workbook_path),
            apply=True,
            allow_reset=False,
            report=str(report_path),
        )
    )

    assert rc == 0
    risks = list((await db_session.execute(select(Risk).order_by(Risk.risk_id_code))).scalars().all())
    assert [risk.risk_id_code for risk in risks] == ["CLAI-R01", "CLAI-R02"]
    assert risks[0].id == existing_risk.id
    assert risks[0].description == "Updated queue narrative"
    assert risks[1].name == "Claims New Label"
    assert risks[1].description == "New alert narrative"

    report = json.loads(report_path.read_text(encoding="utf-8"))
    assert report["creates"] == 1
    assert report["updates"] == 1
    assert report["metadata"]["matched_count"] == 1
    assert report["metadata"]["created_count"] == 1


@pytest.mark.asyncio
async def test_migrate_risks_non_reset_normalizes_process_names_for_department_planning(
    db_session: AsyncSession,
    test_user,
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
):
    workbook_path = tmp_path / "risks-process-casing.xlsx"
    report_path = tmp_path / "risk-process-casing-report.json"
    _write_risk_workbook(
        workbook_path,
        [
            {
                "process": "Claims",
                "subprocess": "Queue",
                "name": "Claims Queue Label",
                "description": "Queue narrative",
            },
            {
                "process": " claims ",
                "subprocess": "Reserve",
                "name": "Claims Reserve Label",
                "description": "Reserve narrative",
            },
        ],
    )

    _patch_script_session(monkeypatch, migrate_risks, db_session)

    rc = await migrate_risks._run(
        argparse.Namespace(
            input=str(workbook_path),
            apply=True,
            allow_reset=False,
            report=str(report_path),
        )
    )

    assert rc == 0
    departments = list((await db_session.execute(select(Department).order_by(Department.id))).scalars().all())
    claims_departments = [department for department in departments if department.name.lower() == "claims"]
    assert len(claims_departments) == 1

    risks = list((await db_session.execute(select(Risk).order_by(Risk.name))).scalars().all())
    assert len(risks) == 2
    assert {risk.department_id for risk in risks} == {claims_departments[0].id}
    assert {risk.process for risk in risks} == {"Claims", "claims"}

    report = json.loads(report_path.read_text(encoding="utf-8"))
    assert report["creates"] == 2
    assert report["updates"] == 0
    assert report["metadata"]["departments_to_create"] == [{"name": "Claims", "code": "CLA"}]


@pytest.mark.asyncio
async def test_migrate_risks_non_reset_reuses_existing_department_when_process_case_differs(
    db_session: AsyncSession,
    test_user,
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
):
    existing_department = Department(name="CLAIMS", code="CLM", description="Claims")
    db_session.add(existing_department)
    await db_session.commit()
    await db_session.refresh(existing_department)

    workbook_path = tmp_path / "risks-existing-department-case.xlsx"
    report_path = tmp_path / "risk-existing-department-case-report.json"
    _write_risk_workbook(
        workbook_path,
        [
            {
                "process": "Claims",
                "subprocess": "Queue",
                "name": "Claims Queue Label",
                "description": "Queue narrative",
            },
            {
                "process": "claims",
                "subprocess": "Reserve",
                "name": "Claims Reserve Label",
                "description": "Reserve narrative",
            },
        ],
    )

    _patch_script_session(monkeypatch, migrate_risks, db_session)

    rc = await migrate_risks._run(
        argparse.Namespace(
            input=str(workbook_path),
            apply=True,
            allow_reset=False,
            report=str(report_path),
        )
    )

    assert rc == 0
    departments = list((await db_session.execute(select(Department).order_by(Department.id))).scalars().all())
    claims_departments = [department for department in departments if department.name.lower() == "claims"]
    assert len(claims_departments) == 1
    assert claims_departments[0].id == existing_department.id

    risks = list((await db_session.execute(select(Risk).order_by(Risk.name))).scalars().all())
    assert len(risks) == 2
    assert {risk.department_id for risk in risks} == {existing_department.id}

    report = json.loads(report_path.read_text(encoding="utf-8"))
    assert report["metadata"]["departments_to_create"] == []


@pytest.mark.asyncio
async def test_migrate_risks_allow_reset_uses_canonical_risk_id_format_and_preserves_series_for_later_non_reset_import(
    db_session: AsyncSession,
    test_user,
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
):
    workbook_path = tmp_path / "risks-reset.xlsx"
    reset_report_path = tmp_path / "risk-reset-report.json"
    _write_risk_workbook(
        workbook_path,
        [
            {
                "process": "Claims",
                "subprocess": "Queue",
                "name": "Claims Queue Label",
                "description": "Queue narrative",
            },
            {
                "process": "Claims",
                "subprocess": "Reserve",
                "name": "Claims Reserve Label",
                "description": "Reserve narrative",
            },
        ],
    )

    _patch_script_session(monkeypatch, migrate_risks, db_session)

    rc = await migrate_risks._run(
        argparse.Namespace(
            input=str(workbook_path),
            apply=True,
            allow_reset=True,
            report=str(reset_report_path),
        )
    )

    assert rc == 0
    reset_risks = list((await db_session.execute(select(Risk).order_by(Risk.risk_id_code))).scalars().all())
    assert [risk.risk_id_code for risk in reset_risks] == ["CLAI-R01", "CLAI-R02"]
    assert not any(risk.risk_id_code.startswith("CLA-R") for risk in reset_risks)

    followup_workbook_path = tmp_path / "risks-followup.xlsx"
    followup_report_path = tmp_path / "risk-followup-report.json"
    _write_risk_workbook(
        followup_workbook_path,
        [
            {
                "process": "Claims",
                "subprocess": "Queue",
                "name": "Claims Queue Label",
                "description": "Queue narrative refreshed",
            },
            {
                "process": "Claims",
                "subprocess": "Reserve",
                "name": "Claims Reserve Label",
                "description": "Reserve narrative refreshed",
            },
            {
                "process": "Claims",
                "subprocess": "Alerts",
                "name": "Claims Alerts Label",
                "description": "Alert narrative",
            },
        ],
    )

    rc = await migrate_risks._run(
        argparse.Namespace(
            input=str(followup_workbook_path),
            apply=True,
            allow_reset=False,
            report=str(followup_report_path),
        )
    )

    assert rc == 0
    followup_risks = list((await db_session.execute(select(Risk).order_by(Risk.risk_id_code))).scalars().all())
    assert [risk.risk_id_code for risk in followup_risks] == ["CLAI-R01", "CLAI-R02", "CLAI-R03"]
    assert followup_risks[2].name == "Claims Alerts Label"
    assert followup_risks[2].description == "Alert narrative"


@pytest.mark.asyncio
async def test_migrate_risks_non_reset_fails_closed_on_ambiguous_existing_matches(
    db_session: AsyncSession,
    test_department: Department,
    test_user,
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
):
    duplicate_a = _risk_fixture(
        risk_id_code="CLAI-R01",
        name="Claims Duplicate Label",
        process="Claims",
        subprocess="Queue",
        description="Duplicate A",
        department_id=test_department.id,
        owner_id=test_user.id,
    )
    duplicate_b = _risk_fixture(
        risk_id_code="CLAI-R02",
        name="Claims Duplicate Label",
        process="Claims",
        subprocess="Queue",
        description="Duplicate B",
        department_id=test_department.id,
        owner_id=test_user.id,
    )
    db_session.add_all([duplicate_a, duplicate_b])
    await db_session.commit()

    workbook_path = tmp_path / "risks-ambiguous.xlsx"
    report_path = tmp_path / "risk-ambiguous-report.json"
    _write_risk_workbook(
        workbook_path,
        [
            {
                "process": "Claims",
                "subprocess": "Queue",
                "name": "Claims Duplicate Label",
                "description": "Replacement narrative",
            }
        ],
    )

    _patch_script_session(monkeypatch, migrate_risks, db_session)

    rc = await migrate_risks._run(
        argparse.Namespace(
            input=str(workbook_path),
            apply=True,
            allow_reset=False,
            report=str(report_path),
        )
    )

    assert rc == 1
    risks = list((await db_session.execute(select(Risk).order_by(Risk.risk_id_code))).scalars().all())
    assert [risk.description for risk in risks] == ["Duplicate A", "Duplicate B"]

    report = json.loads(report_path.read_text(encoding="utf-8"))
    assert report["errors"][0]["code"] == "ambiguous-existing-risk-match"
    assert len(report["metadata"]["ambiguous_matches"]) == 1


@pytest.mark.asyncio
async def test_migrate_risks_fails_closed_on_duplicate_workbook_keys(
    db_session: AsyncSession,
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
):
    workbook_path = tmp_path / "risks-duplicate-keys.xlsx"
    report_path = tmp_path / "risk-duplicate-keys-report.json"
    _write_risk_workbook(
        workbook_path,
        [
            {
                "process": "Claims",
                "subprocess": "Queue",
                "name": "Claims Duplicate Label",
                "description": "First narrative",
            },
            {
                "process": "Claims",
                "subprocess": "Queue",
                "name": "Claims Duplicate Label",
                "description": "Second narrative",
            },
        ],
    )

    _patch_script_session(monkeypatch, migrate_risks, db_session)

    rc = await migrate_risks._run(
        argparse.Namespace(
            input=str(workbook_path),
            apply=True,
            allow_reset=False,
            report=str(report_path),
        )
    )

    assert rc == 1
    risks = list((await db_session.execute(select(Risk))).scalars().all())
    assert risks == []

    report = json.loads(report_path.read_text(encoding="utf-8"))
    assert report["errors"][0]["code"] == "duplicate-workbook-key"
    assert len(report["metadata"]["duplicate_workbook_keys"]) == 1


@pytest.mark.asyncio
async def test_migrate_controls_apply_upserts_existing_control(
    db_session: AsyncSession,
    test_user,
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
):
    ops_department = Department(name="Operations", code="OPS", description="Operations")
    db_session.add(ops_department)
    await db_session.commit()
    await db_session.refresh(ops_department)

    risk = _risk_fixture(
        risk_id_code="FIN-R01",
        name="Finance reconciliation risk",
        process="Finance",
        subprocess=None,
        description="Late reconciliations on core finance process",
        department_id=ops_department.id,
        owner_id=test_user.id,
    )
    existing_control = Control(
        name="Quarterly Finance Review",
        description="Old description",
        control_form="manual",
        frequency="monthly",
        risk_level=3,
        department_id=ops_department.id,
        status="active",
        control_owner_id=test_user.id,
        created_by_id=test_user.id,
        updated_by_id=test_user.id,
    )
    db_session.add_all([risk, existing_control])
    await db_session.commit()
    await db_session.refresh(existing_control)

    workbook_path = tmp_path / "controls.xlsx"
    report_path = tmp_path / "control-report.json"
    _write_controls_workbook(workbook_path)

    _patch_script_session(monkeypatch, migrate_controls, db_session)

    rc = await migrate_controls._run(
        argparse.Namespace(
            input=str(workbook_path),
            apply=True,
            allow_reset=False,
            report=str(report_path),
        )
    )

    assert rc == 0
    controls = list((await db_session.execute(select(Control))).scalars().all())
    assert len(controls) == 1
    assert controls[0].id == existing_control.id
    assert "Updated control description" in controls[0].description
    assert "Escalate repeated breaches" in controls[0].description

    links = list((await db_session.execute(select(ControlRiskLink))).scalars().all())
    assert len(links) == 1
    assert links[0].risk_id == risk.id
    assert links[0].control_id == existing_control.id

    report = json.loads(report_path.read_text(encoding="utf-8"))
    assert report["creates"] == 0
    assert report["updates"] == 1


@pytest.mark.asyncio
async def test_migrate_kris_apply_fails_closed_on_unmatched_rows(
    db_session: AsyncSession,
    test_department: Department,
    test_user,
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
):
    risk = _risk_fixture(
        risk_id_code="OPS-R01",
        name="Existing operational risk",
        process="Operations",
        subprocess=None,
        description="Known matched description",
        department_id=test_department.id,
        owner_id=test_user.id,
    )
    db_session.add(risk)
    await db_session.commit()

    workbook_path = tmp_path / "kris.xlsx"
    report_path = tmp_path / "kri-report.json"
    _write_kri_workbook(workbook_path)

    _patch_script_session(monkeypatch, migrate_kris, db_session)

    rc = await migrate_kris._run(
        argparse.Namespace(
            input=str(workbook_path),
            apply=True,
            allow_reset=False,
            report=str(report_path),
        )
    )

    assert rc == 1
    kris = list((await db_session.execute(select(KeyRiskIndicator))).scalars().all())
    assert kris == []

    report = json.loads(report_path.read_text(encoding="utf-8"))
    assert report["errors"][0]["code"] == "unmatched-kri"
    assert report["creates"] == 0
    assert report["updates"] == 0
